# backend/api/reagents.py
from fastapi import APIRouter, Depends, HTTPException, status, Query, Request, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
from decimal import Decimal
from typing import Optional
from io import BytesIO

from ..db import get_db
from ..schemas.reagents import ReagentCreate, ReagentRead, ReagentBalance
from backend.repositories.reagents_service import (
    create_reagent_supply,
    list_reagent_supplies,
)
from backend.services.reagent_balance_service import ReagentBalanceService
from backend.services.reagent_import_service import ReagentImportService

router = APIRouter(
    prefix="/api/reagents",
    tags=["Reagents"],
)


@router.get("", response_model=list[ReagentRead])
def api_list_reagents(db: Session = Depends(get_db)):
    return list_reagent_supplies(db)


@router.post("", response_model=ReagentRead, status_code=status.HTTP_201_CREATED)
def api_create_reagent(data: ReagentCreate, db: Session = Depends(get_db)):
    dt = data.received_at or datetime.utcnow()
    obj = create_reagent_supply(
        db,
        reagent=data.reagent.strip(),
        qty=data.qty,
        unit=data.unit or "kg",
        received_at=dt,
        source=data.source,
        location=data.location,
        comment=data.comment,
    )
    return obj


@router.get("/supplies")
def api_list_supplies(
    db: Session = Depends(get_db),
    reagent: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    limit: int = Query(100, le=500),
):
    """Список поставок с фильтрами."""
    from backend.models.reagents import ReagentSupply
    from sqlalchemy import desc

    q = db.query(ReagentSupply)
    if reagent:
        q = q.filter(ReagentSupply.reagent == reagent)
    if date_from:
        dt = _parse_date(date_from)
        if dt:
            q = q.filter(ReagentSupply.received_at >= dt)
    if date_to:
        dt = _parse_date(date_to)
        if dt:
            q = q.filter(ReagentSupply.received_at <= dt)

    supplies = q.order_by(desc(ReagentSupply.received_at)).limit(limit).all()
    return [
        {
            "id": s.id,
            "reagent": s.reagent,
            "qty": float(s.qty),
            "unit": s.unit,
            "received_at": s.received_at.isoformat() if s.received_at else None,
            "source": s.source,
            "location": s.location,
            "comment": s.comment,
        }
        for s in supplies
    ]


@router.delete("/supplies/{supply_id}")
def api_delete_supply(
    supply_id: int,
    db: Session = Depends(get_db),
):
    """Удалить поставку по ID."""
    from backend.models.reagents import ReagentSupply

    supply = db.query(ReagentSupply).filter(ReagentSupply.id == supply_id).first()
    if not supply:
        raise HTTPException(status_code=404, detail="Поставку не знайдено")

    info = {"id": supply.id, "reagent": supply.reagent, "qty": float(supply.qty)}
    db.delete(supply)
    db.commit()
    return {"success": True, "deleted": info}


@router.delete("/supplies")
def api_delete_supplies_bulk(
    db: Session = Depends(get_db),
    reagent: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    source: Optional[str] = Query(None),
):
    """Массовое удаление поставок по фильтрам."""
    from backend.models.reagents import ReagentSupply

    q = db.query(ReagentSupply)
    if reagent:
        q = q.filter(ReagentSupply.reagent == reagent)
    if date_from:
        dt = _parse_date(date_from)
        if dt:
            q = q.filter(ReagentSupply.received_at >= dt)
    if date_to:
        dt = _parse_date(date_to)
        if dt:
            q = q.filter(ReagentSupply.received_at <= dt)
    if source:
        q = q.filter(ReagentSupply.source.ilike(f"%{source}%"))

    count = q.count()
    if count == 0:
        return {"success": True, "deleted_count": 0, "message": "Нічого не знайдено"}

    q.delete(synchronize_session=False)
    db.commit()
    return {"success": True, "deleted_count": count}


@router.put("/supplies/{supply_id}")
def api_update_supply(
    supply_id: int,
    db: Session = Depends(get_db),
    reagent: Optional[str] = Query(None),
    qty: Optional[float] = Query(None),
    unit: Optional[str] = Query(None),
    received_at: Optional[str] = Query(None),
    comment: Optional[str] = Query(None),
):
    """Редактировать поставку."""
    from backend.models.reagents import ReagentSupply

    supply = db.query(ReagentSupply).filter(ReagentSupply.id == supply_id).first()
    if not supply:
        raise HTTPException(status_code=404, detail="Поставку не знайдено")

    if reagent is not None:
        supply.reagent = reagent
    if qty is not None:
        supply.qty = qty
    if unit is not None:
        supply.unit = unit
    if received_at is not None:
        dt = _parse_date(received_at)
        if dt:
            supply.received_at = dt
    if comment is not None:
        supply.comment = comment

    db.commit()
    return {
        "success": True,
        "supply": {
            "id": supply.id,
            "reagent": supply.reagent,
            "qty": float(supply.qty),
            "unit": supply.unit,
            "received_at": supply.received_at.isoformat() if supply.received_at else None,
        }
    }


@router.get("/balance")
def api_reagents_balance(
    db: Session = Depends(get_db),
    as_of: Optional[str] = Query(None, description="Дата расчёта (YYYY-MM-DD или ISO)")
):
    """
    Возвращает актуальные остатки реагентов с учётом:
    - Приход (reagent_supplies)
    - Расход (events где event_type='reagent')
    - Инвентаризация (reagent_inventory)

    Формула:
    - Если есть инвентаризация до даты: факт + приход_после - расход_после
    - Иначе: весь_приход - весь_расход
    """
    as_of_date = None
    if as_of:
        try:
            as_of_date = datetime.fromisoformat(as_of.replace("Z", "+00:00"))
        except ValueError:
            try:
                as_of_date = datetime.strptime(as_of, "%Y-%m-%d")
            except ValueError:
                pass

    balances = ReagentBalanceService.get_all_reagents_balance(db, as_of_date)

    # Преобразуем в формат API
    result = []
    for b in balances:
        result.append({
            "reagent": b["reagent"],
            "current_balance": float(b["current_balance"]),
            "calculation_method": b["calculation_method"],
            "as_of_date": b["as_of_date"].isoformat() if b.get("as_of_date") else None,
            "last_inventory_date": b.get("last_inventory_date").isoformat() if b.get("last_inventory_date") else None,
        })

    return result


@router.get("/balance/debug/{reagent_name}")
def api_reagent_balance_debug(
    reagent_name: str,
    db: Session = Depends(get_db),
):
    """
    Диагностика расчёта остатка конкретного реагента.
    Показывает все составляющие: инвентаризации, поставки, расход.
    """
    from backend.models.reagent_inventory import ReagentInventorySnapshot
    from backend.models.reagents import ReagentSupply
    from backend.models.events import Event
    from sqlalchemy import desc

    balance = ReagentBalanceService.get_current_balance(db, reagent_name)

    # Все инвентаризации
    inventories = (
        db.query(ReagentInventorySnapshot)
        .filter(ReagentInventorySnapshot.reagent == reagent_name)
        .order_by(desc(ReagentInventorySnapshot.snapshot_at))
        .limit(10)
        .all()
    )

    # Все поставки
    supplies = (
        db.query(ReagentSupply)
        .filter(ReagentSupply.reagent == reagent_name)
        .order_by(desc(ReagentSupply.received_at))
        .limit(20)
        .all()
    )

    # Последние события расхода
    events = (
        db.query(Event)
        .filter(Event.reagent == reagent_name, Event.event_type == "reagent")
        .order_by(desc(Event.event_time))
        .limit(20)
        .all()
    )

    return {
        "reagent": reagent_name,
        "balance": {k: str(v) if isinstance(v, (Decimal,)) else (v.isoformat() if isinstance(v, datetime) else v)
                    for k, v in balance.items()},
        "inventories": [
            {"id": i.id, "qty": str(i.qty), "snapshot_at": i.snapshot_at.isoformat() if i.snapshot_at else None,
             "calculated_qty": str(i.calculated_qty) if i.calculated_qty else None,
             "discrepancy": str(i.discrepancy) if i.discrepancy else None}
            for i in inventories
        ],
        "supplies": [
            {"id": s.id, "qty": str(s.qty), "received_at": s.received_at.isoformat() if s.received_at else None,
             "source": s.source, "comment": s.comment}
            for s in supplies
        ],
        "consumption_events": [
            {"id": e.id, "qty": str(e.qty), "event_time": e.event_time.isoformat() if e.event_time else None,
             "well": e.well}
            for e in events
        ],
    }


# =============================================================================
# АНАЛИТИКА
# =============================================================================

def _parse_date(date_str: Optional[str]) -> Optional[datetime]:
    """Парсит дату из строки."""
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str.replace("Z", "+00:00"))
    except ValueError:
        try:
            return datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            return None


@router.get("/analytics/daily-consumption")
def api_daily_consumption(
    db: Session = Depends(get_db),
    date_from: str = Query(..., description="Начало периода (YYYY-MM-DD)"),
    date_to: str = Query(..., description="Конец периода (YYYY-MM-DD)"),
    reagents: Optional[str] = Query(None, description="Реагенты через запятую"),
    wells: Optional[str] = Query(None, description="Скважины через запятую"),
):
    """
    Расход реагентов по дням за период.
    Группировка по дате и реагенту.
    """
    from_dt = _parse_date(date_from)
    to_dt = _parse_date(date_to)

    if not from_dt or not to_dt:
        return {"error": "Invalid date format"}

    reagent_list = [r.strip() for r in reagents.split(",")] if reagents else None
    well_list = [w.strip() for w in wells.split(",")] if wells else None

    data = ReagentBalanceService.get_daily_consumption(
        db, from_dt, to_dt, reagent_list, well_list
    )
    return {"data": data, "date_from": date_from, "date_to": date_to}


@router.get("/analytics/consumption-by-wells")
def api_consumption_by_wells(
    db: Session = Depends(get_db),
    date_from: str = Query(..., description="Начало периода (YYYY-MM-DD)"),
    date_to: str = Query(..., description="Конец периода (YYYY-MM-DD)"),
    reagents: Optional[str] = Query(None, description="Реагенты через запятую"),
    wells: Optional[str] = Query(None, description="Скважины через запятую"),
    statuses: Optional[str] = Query(None, description="Статуси скважин через запятую"),
):
    """
    Расход реагентов по скважинам за период.
    Поддерживает фильтр по скважинам, реагентам и статусам.
    """
    from_dt = _parse_date(date_from)
    to_dt = _parse_date(date_to)

    if not from_dt or not to_dt:
        return {"error": "Invalid date format"}

    reagent_list = [r.strip() for r in reagents.split(",")] if reagents else None
    well_list = [w.strip() for w in wells.split(",")] if wells else None
    status_list = [s.strip() for s in statuses.split(",")] if statuses else None

    data = ReagentBalanceService.get_consumption_by_wells(
        db, from_dt, to_dt, reagent_list, well_list, status_list
    )
    return {"data": data, "date_from": date_from, "date_to": date_to}


@router.get("/analytics/reagent-stats/{reagent_name}")
def api_reagent_statistics(
    reagent_name: str,
    db: Session = Depends(get_db),
    date_from: str = Query(..., description="Начало периода (YYYY-MM-DD)"),
    date_to: str = Query(..., description="Конец периода (YYYY-MM-DD)"),
):
    """
    Детальная статистика по одному реагенту за период.
    """
    from_dt = _parse_date(date_from)
    to_dt = _parse_date(date_to)

    if not from_dt or not to_dt:
        return {"error": "Invalid date format"}

    data = ReagentBalanceService.get_reagent_statistics(
        db, reagent_name, from_dt, to_dt
    )
    return data


@router.get("/analytics/forecast")
def api_depletion_forecast(
    db: Session = Depends(get_db),
    avg_days: int = Query(30, description="Дней для расчёта среднего расхода"),
):
    """
    Прогноз даты окончания для всех реагентов.
    """
    forecasts = ReagentBalanceService.get_all_depletion_forecasts(db, avg_days)
    return {"data": forecasts, "calculation_days": avg_days}


@router.get("/analytics/forecast-by-wells")
def api_forecast_by_wells(
    db: Session = Depends(get_db),
    avg_days: int = Query(30, description="Дней для расчёта среднего расхода"),
):
    """
    Прогноз вичерпання реагентів, згрупований по скважинах.

    Для кожної скважини:
    - Список реагентів, що використовуються
    - Залишок (загальний складський)
    - Середній розхід на цій скважині
    - Прогноз днів до вичерпання
    """
    forecasts = ReagentBalanceService.get_forecast_by_wells(db, avg_days)
    return {"data": forecasts, "calculation_days": avg_days}


@router.get("/analytics/forecast-by-reagents")
def api_forecast_by_reagents(
    db: Session = Depends(get_db),
    date_from: Optional[str] = Query(None, description="Начало периода (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Конец периода (YYYY-MM-DD)"),
    reagents: Optional[str] = Query(None, description="Реагенти через кому"),
    wells: Optional[str] = Query(None, description="Скважини через кому"),
    statuses: Optional[str] = Query(None, description="Статуси скважин через кому"),
    lead_days: int = Query(15, description="Днів до закінчення для рекомендації замовлення"),
):
    """
    Прогноз вичерпання реагентів (складський рівень) з фільтрами.

    Повертає для кожного реагента:
    - Поточний залишок
    - Середній добовий розхід (за вибраний період)
    - Днів до вичерпання
    - Дату вичерпання
    - Дату "замовити до" (вичерпання - lead_days)
    - Статус ризику (critical/warning/normal/no_data)
    - ТОП споживачів (скважини)
    """
    from_dt = _parse_date(date_from)
    to_dt = _parse_date(date_to)

    reagent_list = [r.strip() for r in reagents.split(",")] if reagents else None
    well_list = [w.strip() for w in wells.split(",")] if wells else None
    status_list = [s.strip() for s in statuses.split(",")] if statuses else None

    forecasts = ReagentBalanceService.get_reagent_forecasts_with_filters(
        db=db,
        date_from=from_dt,
        date_to=to_dt,
        reagent_names=reagent_list,
        well_numbers=well_list,
        well_statuses=status_list,
        lead_days=lead_days,
    )

    return {
        "data": forecasts,
        "params": {
            "date_from": date_from,
            "date_to": date_to,
            "lead_days": lead_days,
        }
    }


@router.get("/analytics/procurement-plan")
def api_procurement_plan(
    db: Session = Depends(get_db),
    target_days: int = Query(60, description="На скільки днів забезпечити"),
    lead_days: int = Query(15, description="Днів до закінчення для рекомендації"),
    date_from: Optional[str] = Query(None, description="Початок періоду для avg_daily"),
    date_to: Optional[str] = Query(None, description="Кінець періоду для avg_daily"),
    reagents: Optional[str] = Query(None, description="Реагенти через кому"),
    wells: Optional[str] = Query(None, description="Скважини через кому"),
    statuses: Optional[str] = Query(None, description="Статуси скважин через кому"),
    include_depleted: bool = Query(False, description="Включити реагенти з залишком 0"),
):
    """
    План закупки реагентів.

    Розрахунок:
    - needed_total = avg_daily * target_days
    - to_order = max(0, needed_total - stock_now)

    Повертає таблицю рекомендацій з пріоритетами.
    """
    from_dt = _parse_date(date_from)
    to_dt = _parse_date(date_to)

    reagent_list = [r.strip() for r in reagents.split(",")] if reagents else None
    well_list = [w.strip() for w in wells.split(",")] if wells else None
    status_list = [s.strip() for s in statuses.split(",")] if statuses else None

    plan = ReagentBalanceService.calculate_procurement_plan(
        db=db,
        target_days=target_days,
        lead_days=lead_days,
        date_from=from_dt,
        date_to=to_dt,
        reagent_names=reagent_list,
        well_numbers=well_list,
        well_statuses=status_list,
        include_depleted=include_depleted,
    )

    return plan


@router.get("/analytics/procurement-plan/export")
def api_procurement_plan_export(
    db: Session = Depends(get_db),
    target_days: int = Query(60, description="На скільки днів забезпечити"),
    lead_days: int = Query(15, description="Днів до закінчення для рекомендації"),
    date_from: Optional[str] = Query(None, description="Початок періоду для avg_daily"),
    date_to: Optional[str] = Query(None, description="Кінець періоду для avg_daily"),
    reagents: Optional[str] = Query(None, description="Реагенти через кому"),
    wells: Optional[str] = Query(None, description="Скважини через кому"),
    statuses: Optional[str] = Query(None, description="Статуси скважин через кому"),
    include_depleted: bool = Query(False, description="Включити реагенти з залишком 0"),
):
    """
    Експорт плану закупки в Excel.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    from_dt = _parse_date(date_from)
    to_dt = _parse_date(date_to)

    reagent_list = [r.strip() for r in reagents.split(",")] if reagents else None
    well_list = [w.strip() for w in wells.split(",")] if wells else None
    status_list = [s.strip() for s in statuses.split(",")] if statuses else None

    plan = ReagentBalanceService.calculate_procurement_plan(
        db=db,
        target_days=target_days,
        lead_days=lead_days,
        date_from=from_dt,
        date_to=to_dt,
        reagent_names=reagent_list,
        well_numbers=well_list,
        well_statuses=status_list,
        include_depleted=include_depleted,
    )

    # Створюємо Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "План закупки"

    # Стилі
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовок документа
    ws.merge_cells('A1:H1')
    ws['A1'] = f"План закупки реагентів на {target_days} днів"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Параметри розрахунку
    ws['A3'] = "Параметри розрахунку:"
    ws['A3'].font = header_font
    ws['A4'] = f"Горизонт планування: {target_days} днів"
    ws['A5'] = f"Lead time: {lead_days} днів"
    ws['A6'] = f"База розрахунку: {plan['params']['period_days']} днів ({plan['params']['date_from']} - {plan['params']['date_to']})"
    ws['A7'] = f"Дата формування: {plan['params']['calculated_at']}"

    # Підсумки
    ws['A9'] = "Підсумки:"
    ws['A9'].font = header_font
    ws['A10'] = f"Всього реагентів: {plan['summary']['total_items']}"
    ws['A11'] = f"Потребують замовлення: {plan['summary']['items_to_order']}"
    ws['A12'] = f"Критичних: {plan['summary']['critical_count']}"
    ws['A13'] = f"Попередження: {plan['summary']['warning_count']}"

    # Таблиця
    row = 15
    headers = ["Реагент", "Од.", "Залишок", "Сер./день", "Потрібно", "Замовити", "Замовити до", "Пріоритет"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Дані
    priority_colors = {
        "high": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "medium": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "low": None,
    }
    priority_labels = {"high": "Високий", "medium": "Середній", "low": "Низький"}

    for item in plan['items']:
        row += 1
        ws.cell(row=row, column=1, value=item['reagent']).border = thin_border
        ws.cell(row=row, column=2, value=item['unit']).border = thin_border
        ws.cell(row=row, column=3, value=item['stock']).border = thin_border
        ws.cell(row=row, column=4, value=item['avg_daily']).border = thin_border
        ws.cell(row=row, column=5, value=item['needed_total']).border = thin_border
        ws.cell(row=row, column=6, value=item['to_order']).border = thin_border
        ws.cell(row=row, column=7, value=item['order_by_date'] or "—").border = thin_border
        ws.cell(row=row, column=8, value=priority_labels.get(item['priority'], "—")).border = thin_border

        # Підсвічування рядка по пріоритету
        fill = priority_colors.get(item['priority'])
        if fill:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = fill

    # Ширина колонок
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12

    # Зберігаємо у буфер
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"procurement_plan_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/analytics/filters")
def api_analytics_filters(
    db: Session = Depends(get_db),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
):
    """
    Возвращает доступные значения для фильтров аналитики:
    - Уникальные реагенты
    - Уникальные скважины
    - Доступные статусы скважин
    """
    from backend.models.well_status import ALLOWED_STATUS

    from_dt = _parse_date(date_from)
    to_dt = _parse_date(date_to)

    reagents = ReagentBalanceService.get_unique_reagents_from_events(db, from_dt, to_dt)
    wells = ReagentBalanceService.get_unique_wells_from_events(db, from_dt, to_dt)

    return {
        "reagents": reagents,
        "wells": wells,
        "statuses": list(ALLOWED_STATUS),
    }


# =============================================================================
# ИМПОРТ ИЗ EXCEL
# =============================================================================

@router.get("/import/template/supplies")
def download_supply_template(db: Session = Depends(get_db)):
    """Скачать шаблон Excel для импорта поступлений."""
    content = ReagentImportService.generate_supply_template(db)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=reagent_supplies_template.xlsx"
        }
    )


@router.get("/import/template/inventory")
def download_inventory_template(db: Session = Depends(get_db)):
    """Скачать шаблон Excel для импорта инвентаризаций."""
    content = ReagentImportService.generate_inventory_template(db)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=reagent_inventory_template.xlsx"
        }
    )


@router.post("/import/supplies/validate")
async def validate_supply_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Валидация файла поступлений (dry-run).
    Возвращает список валидных строк и ошибок без сохранения в БД.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        return {"success": False, "message": "Файл повинен бути у форматі Excel (.xlsx)"}

    content = await file.read()
    result = ReagentImportService.validate_supply_file(db, content)

    # Список каталожных реагентов для JS (dropdown при ошибках)
    from backend.models.reagent_catalog import ReagentCatalog
    catalog_items = db.query(ReagentCatalog).filter(
        ReagentCatalog.is_active == True  # noqa: E712
    ).order_by(ReagentCatalog.name).all()
    catalog_list = [{"name": c.name, "unit": c.default_unit or "шт"} for c in catalog_items]

    return {
        "success": result.success,
        "message": result.message,
        "total_rows": result.total_rows,
        "valid_count": len(result.valid_rows),
        "error_count": len(result.errors),
        "valid_rows": [
            {
                "row": r.row_num,
                "date": r.date.strftime("%d.%m.%Y") if r.date else None,
                "reagent": r.reagent,
                "qty": float(r.qty) if r.qty else None,
                "unit": r.unit,
                "comment": r.comment,
            }
            for r in result.valid_rows
        ],
        "errors": [
            {
                "row": e.row,
                "column": e.column,
                "message": e.message,
                "partial_data": e.partial_data,
            }
            for e in result.errors
        ],
        "catalog": catalog_list,
    }


@router.post("/import/supplies")
async def import_supplies(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    created_by: Optional[str] = Form(None),
):
    """
    Импорт поступлений из Excel файла.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        return {"success": False, "message": "Файл повинен бути у форматі Excel (.xlsx)"}

    content = await file.read()
    result = ReagentImportService.import_supplies(db, content, created_by)

    return {
        "success": result.success,
        "message": result.message,
        "imported_count": result.imported_count,
        "error_count": len(result.errors),
        "errors": [
            {"row": e.row, "column": e.column, "message": e.message}
            for e in result.errors
        ],
    }


@router.post("/import/supplies/json")
async def import_supplies_json(
    request: Request,
    db: Session = Depends(get_db),
):
    """
    Импорт поставок из JSON (отредактированные строки из preview).
    Body: {"rows": [{"date": "20.03.2026", "reagent": "...", "qty": 90, "unit": "шт", "comment": "..."}, ...]}
    """
    from backend.models.reagents import ReagentSupply
    from backend.models.reagent_catalog import ReagentCatalog

    body = await request.json()
    rows = body.get("rows", [])
    if not rows:
        return {"success": False, "message": "Немає рядків для імпорту"}

    catalog_items = db.query(ReagentCatalog).filter(ReagentCatalog.is_active == True).all()
    catalog = {r.name: r for r in catalog_items}
    catalog_lower = {r.name.lower(): r for r in catalog_items}

    imported = 0
    created_catalog = []
    errors = []
    for i, row in enumerate(rows):
        reagent_raw = (row.get("reagent") or "").strip()
        if not reagent_raw:
            errors.append({"row": i + 1, "message": "Реагент не вказано"})
            continue

        cat = catalog.get(reagent_raw) or catalog_lower.get(reagent_raw.lower())
        if not cat:
            # Автосоздание нового реагента в каталоге
            new_cat = ReagentCatalog(
                name=reagent_raw,
                default_unit=row.get("unit") or "шт",
                is_active=True,
            )
            db.add(new_cat)
            db.flush()  # получаем id
            cat = new_cat
            catalog[cat.name] = cat
            catalog_lower[cat.name.lower()] = cat
            created_catalog.append(cat.name)

        try:
            qty = float(str(row.get("qty", 0)).replace(",", "."))
        except (ValueError, TypeError):
            errors.append({"row": i + 1, "message": f"Невірна кількість: {row.get('qty')}"})
            continue

        if qty <= 0:
            errors.append({"row": i + 1, "message": f"Кількість повинна бути > 0"})
            continue

        date_str = row.get("date", "")
        dt = None
        for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
            try:
                dt = datetime.strptime(date_str, fmt)
                break
            except (ValueError, TypeError):
                continue
        if not dt:
            try:
                dt = datetime.fromisoformat(date_str)
            except (ValueError, TypeError):
                dt = datetime.now()
        # Если дата без времени (полночь) — ставим 23:59:59,
        # чтобы поставка всегда была "после" инвентаризации/расхода в тот же день
        if dt and dt.hour == 0 and dt.minute == 0 and dt.second == 0:
            dt = dt.replace(hour=23, minute=59, second=59)

        supply = ReagentSupply(
            reagent=cat.name,
            reagent_id=cat.id,
            qty=qty,
            unit=row.get("unit") or cat.default_unit or "шт",
            received_at=dt,
            source="Excel import (edited)",
            comment=row.get("comment") or None,
        )
        db.add(supply)
        imported += 1

    if imported > 0:
        db.commit()

    msg = f"Імпортовано {imported} записів" if imported > 0 else "Нічого не імпортовано"
    if created_catalog:
        msg += f". Створено нових реагентів: {', '.join(created_catalog)}"

    return {
        "success": imported > 0,
        "imported_count": imported,
        "error_count": len(errors),
        "errors": errors,
        "created_catalog": created_catalog,
        "message": msg,
    }


@router.post("/import/inventory/validate")
async def validate_inventory_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Валидация файла инвентаризации (dry-run).
    Показывает расчётные остатки и расхождения.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        return {"success": False, "message": "Файл повинен бути у форматі Excel (.xlsx)"}

    content = await file.read()
    result = ReagentImportService.validate_inventory_file(db, content)

    return {
        "success": result.success,
        "message": result.message,
        "total_rows": result.total_rows,
        "valid_count": len(result.valid_rows),
        "error_count": len(result.errors),
        "valid_rows": [
            {
                "row": r.row_num,
                "date": r.date.strftime("%d.%m.%Y") if r.date else None,
                "reagent": r.reagent,
                "actual_qty": float(r.qty) if r.qty else None,
                "calculated_qty": float(r.calculated_qty) if r.calculated_qty is not None else None,
                "discrepancy": float(r.discrepancy) if r.discrepancy is not None else None,
                "comment": r.comment,
            }
            for r in result.valid_rows
        ],
        "errors": [
            {"row": e.row, "column": e.column, "message": e.message}
            for e in result.errors
        ],
    }


@router.post("/import/inventory")
async def import_inventory(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    created_by: Optional[str] = Form(None),
):
    """
    Импорт инвентаризации из Excel файла.
    Автоматически рассчитывает calculated_qty и discrepancy.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        return {"success": False, "message": "Файл повинен бути у форматі Excel (.xlsx)"}

    content = await file.read()
    result = ReagentImportService.import_inventory(db, content, created_by)

    return {
        "success": result.success,
        "message": result.message,
        "imported_count": result.imported_count,
        "error_count": len(result.errors),
        "errors": [
            {"row": e.row, "column": e.column, "message": e.message}
            for e in result.errors
        ],
    }