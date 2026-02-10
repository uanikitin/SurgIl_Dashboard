"""
Роутер для управління обладнанням
Сторінка перегляду, фільтрації, переміщення обладнання
"""

from __future__ import annotations

from datetime import datetime, timedelta, date
from typing import List, Optional
from fastapi import APIRouter, Depends, Form, HTTPException, Request, Query
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, desc, func
# from starlette.responses import RedirectResponse

from backend.db import get_db
from backend.web.templates import templates
from backend.deps import get_current_admin
from backend.models.wells import Well
from backend.models.equipment import Equipment, EquipmentInstallation, EquipmentMaintenance
from backend.models.users import User
from backend.documents.models import Document
import io


router = APIRouter(tags=["equipment-management"])


# ======================================================================================
# Форма додавання обладнання
# ======================================================================================

@router.get("/equipment/add", response_class=HTMLResponse)
async def equipment_add_page(
    request: Request,
    db: Session = Depends(get_db),
):
    """Форма додавання нового обладнання"""

    context = {
        "request": request,
    }

    return templates.TemplateResponse("equipment_add.html", context)


# ======================================================================================
# API: Створення обладнання
# ======================================================================================

@router.post("/api/equipment/create")
async def create_equipment(
    name: str = Form(...),
    equipment_type: str = Form(...),
    serial_number: str = Form(...),
    manufacturer: Optional[str] = Form(None),
    manufacture_date: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    specifications: Optional[str] = Form(None),
    current_location: str = Form("Склад"),
    condition: str = Form("working"),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_admin = Depends(get_current_admin),
):
    """
    Створення нового обладнання
    """

    # Перевірка унікальності серійного номера
    existing = db.query(Equipment).filter(
        Equipment.serial_number == serial_number
    ).first()

    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Обладнання з серійним номером {serial_number} вже існує (ID: {existing.id})"
        )

    # Парсимо дату виробництва
    manufacture_dt = None
    if manufacture_date:
        try:
            from datetime import datetime as dt
            manufacture_dt = dt.fromisoformat(manufacture_date)
        except:
            pass

    # Створюємо обладнання
    equipment = Equipment(
        name=name,
        equipment_type=equipment_type,
        serial_number=serial_number,
        manufacturer=manufacturer,
        manufacture_date=manufacture_dt,
        description=description,
        specifications=specifications,
        status='available',  # Завжди створюється як "на складі"
        condition=condition,
        current_location=current_location,
        notes=notes,
        created_at=datetime.now(),
        updated_at=datetime.now(),
    )

    db.add(equipment)
    db.commit()
    db.refresh(equipment)

    return RedirectResponse(
        url=f"/equipment/view/{equipment.id}",
        status_code=303
    )


# ======================================================================================
# Список обладнання
# ======================================================================================

@router.get("/equipment", response_class=HTMLResponse)
async def equipment_list_page(
    request: Request,
    db: Session = Depends(get_db),
    equipment_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    condition: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
):
    """
    Сторінка списку обладнання з фільтрами

    Фільтри:
    - equipment_type: тип обладнання (шлюз, манометр, etc)
    - status: available, installed, maintenance, broken, lost
    - condition: working, broken, needs_repair
    - search: пошук по назві або серійному номеру
    """

    # Базовий запит
    query = db.query(Equipment).filter(
        or_(Equipment.deleted_at.is_(None), Equipment.deleted_at == None)
    )

    # Фільтр по типу
    if equipment_type:
        query = query.filter(Equipment.equipment_type == equipment_type)

    # Фільтр по статусу
    if status:
        query = query.filter(Equipment.status == status)

    # Фільтр по стану
    if condition:
        query = query.filter(Equipment.condition == condition)

    # Пошук
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            or_(
                Equipment.name.ilike(search_pattern),
                Equipment.serial_number.ilike(search_pattern),
            )
        )

    # Сортування
    query = query.order_by(Equipment.id.desc())

    # Отримуємо дані
    equipment_list = query.all()

    # Отримуємо унікальні типи для фільтру
    equipment_types = db.query(Equipment.equipment_type).distinct().filter(
        Equipment.equipment_type.isnot(None)
    ).all()
    equipment_types = [t[0] for t in equipment_types if t[0]]

    # Статистика
    total_count = db.query(func.count(Equipment.id)).scalar()
    installed_count = db.query(func.count(Equipment.id)).filter(
        Equipment.status == 'installed'
    ).scalar()
    available_count = db.query(func.count(Equipment.id)).filter(
        Equipment.status == 'available'
    ).scalar()
    maintenance_count = db.query(func.count(Equipment.id)).filter(
        Equipment.status == 'maintenance'
    ).scalar()

    context = {
        "request": request,
        "equipment_list": equipment_list,
        "equipment_types": equipment_types,
        "selected_type": equipment_type,
        "selected_status": status,
        "selected_condition": condition,
        "search_query": search or "",
        "stats": {
            "total": total_count,
            "installed": installed_count,
            "available": available_count,
            "maintenance": maintenance_count,
        }
    }

    return templates.TemplateResponse("equipment_list.html", context)


# ======================================================================================
# Деталі обладнання
# ======================================================================================

@router.get("/equipment/view/{equipment_id}", response_class=HTMLResponse)
async def equipment_detail_page(
    request: Request,
    equipment_id: int,
    db: Session = Depends(get_db),
):
    """
    Детальна інформація про обладнання:
    - Поточний стан
    - Історія встановлень
    - Історія обслуговування
    - Статистика
    """

    # Отримуємо обладнання
    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        raise HTTPException(status_code=404, detail="Обладнання не знайдено")

    # Історія встановлень (з інформацією про свердловини та користувачів)
    installations_query = db.query(
        EquipmentInstallation,
        Well.number.label('well_number')
    ).outerjoin(
        Well, EquipmentInstallation.well_id == Well.id
    ).filter(
        EquipmentInstallation.equipment_id == equipment_id
    ).order_by(
        EquipmentInstallation.installed_at.desc()
    )

    installations_raw = installations_query.all()

    # Отримуємо унікальні user_id для завантаження
    user_ids = set()
    for inst, _ in installations_raw:
        if inst.installed_by:
            try:
                user_ids.add(int(inst.installed_by))
            except (ValueError, TypeError):
                pass
        if inst.removed_by:
            try:
                user_ids.add(int(inst.removed_by))
            except (ValueError, TypeError):
                pass

    # Завантажуємо users одним запитом
    users_dict = {}
    if user_ids:
        users = db.query(User).filter(User.id.in_(user_ids)).all()
        for user in users:
            users_dict[user.id] = user.username or user.full_name or f"User {user.id}"

    # Обробка історії
    installation_history = []
    for inst, well_num in installations_raw:
        start_dt = inst.installed_at
        end_dt = inst.removed_at or datetime.now()
        duration_days = (end_dt - start_dt).days

        installed_by_name = None
        if inst.installed_by:
            try:
                installed_by_name = users_dict.get(int(inst.installed_by))
            except (ValueError, TypeError):
                pass

        removed_by_name = None
        if inst.removed_by:
            try:
                removed_by_name = users_dict.get(int(inst.removed_by))
            except (ValueError, TypeError):
                pass

        has_document = inst.document_id is not None
        no_document_confirmed = getattr(inst, 'no_document_confirmed', False) or False

        installation_history.append({
            "id": inst.id,
            "well_number": well_num,
            "well_id": inst.well_id,
            "installed_at": inst.installed_at,
            "removed_at": inst.removed_at,
            "is_active": inst.removed_at is None,
            "duration_days": duration_days,
            "installed_by": installed_by_name or (f"User {inst.installed_by}" if inst.installed_by else None),
            "removed_by": removed_by_name or (f"User {inst.removed_by}" if inst.removed_by else None),
            "installation_location": getattr(inst, 'installation_location', None),
            "has_document": has_document,
            "no_document_confirmed": no_document_confirmed,
            "document_id": inst.document_id,
            "notes": inst.notes,
        })

    # Історія обслуговування
    maintenance_history = db.query(EquipmentMaintenance).filter(
        EquipmentMaintenance.equipment_id == equipment_id
    ).order_by(
        EquipmentMaintenance.maintenance_date.desc()
    ).all()

    # Статистика
    total_installations = len(installation_history)
    total_days = sum(h['duration_days'] for h in installation_history)
    total_maintenance_cost = sum(m.cost or 0 for m in maintenance_history)

    # Поточна установка
    current_installation = next((h for h in installation_history if h['is_active']), None)

    # Додаємо well_id для створення акту демонтажу
    if current_installation:
        for inst, _well_num in installations_raw:
            if inst.removed_at is None:
                current_installation['well_id'] = inst.well_id
                break

    # Список свердловин для форми встановлення
    wells_list = db.query(Well).order_by(Well.number).all()

    context = {
        "request": request,
        "equipment": equipment,
        "current_installation": current_installation,
        "installation_history": installation_history,
        "maintenance_history": maintenance_history,
        "wells_list": wells_list,
        "stats": {
            "total_installations": total_installations,
            "total_days": total_days,
            "total_maintenance_cost": total_maintenance_cost,
            "avg_days_per_installation": round(total_days / total_installations, 1) if total_installations > 0 else 0,
        }
    }

    return templates.TemplateResponse("equipment_detail.html", context)


# ======================================================================================
# API: Переміщення обладнання
# ======================================================================================

@router.post("/api/equipment/{equipment_id}/move")
async def move_equipment(
    equipment_id: int,
    action: str = Form(...),  # "install" або "remove"
    well_id: Optional[str] = Form(None),  # ЗМІНЕНО: строка, потім конвертуємо
    installation_location: Optional[str] = Form(None),  # НОВЕ: Устье, НКТ, Шлейф, Затруб, Інше
    location: Optional[str] = Form(None),
    condition: str = Form("working"),
    create_document: bool = Form(False),
    no_document_confirmed: bool = Form(False),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_admin = Depends(get_current_admin),
):
    """
    API для переміщення обладнання

    Parameters:
    - action: "install" (встановити) або "remove" (зняти)
    - well_id: ID свердловини (обов'язково для install)
    - location: Локація текстом (для remove, якщо не "Склад")
    - condition: "working", "broken", "needs_repair"
    - create_document: чи створювати акт
    - no_document_confirmed: підтвердження що акт не потрібен
    - notes: примітки
    """

    # Отримуємо обладнання
    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        raise HTTPException(status_code=404, detail="Обладнання не знайдено")

    # Конвертуємо well_id з строки в int
    well_id_int = None
    if well_id and well_id.strip():
        try:
            well_id_int = int(well_id)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Невалідний well_id: '{well_id}'")

    # Перевіряємо поточний стан
    active_installation = db.query(EquipmentInstallation).filter(
        EquipmentInstallation.equipment_id == equipment_id,
        EquipmentInstallation.removed_at.is_(None)
    ).first()

    if action == "install":
        # ==== ВСТАНОВЛЕННЯ ====

        # Якщо є активна установка на ІНШІЙ свердловині — автоматично закриваємо
        if active_installation:
            if active_installation.well_id == well_id_int:
                raise HTTPException(
                    status_code=400,
                    detail=f"Обладнання вже встановлено на цій свердловині"
                )
            # Автозакриття старої установки (дата = дата нової установки)
            active_installation.removed_at = datetime.now()
            active_installation.removed_by = str(current_admin.id)
            active_installation.notes = (
                (active_installation.notes or "")
                + f"\nАвтозакриття: обладнання встановлено на іншу свердловину"
            )

        # Перевірка: well_id обов'язковий
        if not well_id_int:
            raise HTTPException(status_code=400, detail="Виберіть свердловину для встановлення")

        # Перевіряємо що свердловина існує
        well = db.query(Well).filter(Well.id == well_id_int).first()
        if not well:
            raise HTTPException(status_code=404, detail=f"Свердловина {well_id_int} не знайдена")

        # Створюємо запис установки
        installation = EquipmentInstallation(
            equipment_id=equipment_id,
            well_id=well_id_int,
            installed_at=datetime.now(),
            removed_at=None,
            tube_pressure_install=None,
            line_pressure_install=None,
            installed_by=str(current_admin.id),
            installation_location=installation_location,
            document_id=None,
            no_document_confirmed=no_document_confirmed,
            notes=notes,
        )
        db.add(installation)

        # Оновлюємо обладнання
        equipment.status = 'installed'
        equipment.current_location = f"Свердловина {well.number}"
        equipment.condition = condition

        db.commit()

        return JSONResponse({
            "success": True,
            "message": f"Обладнання встановлено на свердловині {well.number}",
            "installation_id": installation.id,
            "requires_document": not no_document_confirmed,
        })

    elif action == "remove":
        # ==== ДЕМОНТАЖ ====

        # Перевірка: чи встановлено зараз
        if not active_installation:
            raise HTTPException(
                status_code=400,
                detail="Обладнання не встановлено"
            )

        # Закриваємо установку
        active_installation.removed_at = datetime.now()
        active_installation.removed_by = current_admin.id
        if notes:
            active_installation.notes = (active_installation.notes or "") + f"\nДемонтаж: {notes}"

        # Оновлюємо обладнання
        equipment.status = 'available'
        equipment.current_location = location or "Склад"
        equipment.condition = condition

        db.commit()

        return JSONResponse({
            "success": True,
            "message": f"Обладнання знято. Локація: {equipment.current_location}",
            "installation_id": active_installation.id,
        })

    else:
        raise HTTPException(status_code=400, detail="action має бути 'install' або 'remove'")


# ======================================================================================
# API: Переміщення датчика на іншу свердловину (атомарний transfer)
# ======================================================================================

@router.post("/api/equipment/{equipment_id}/transfer")
async def transfer_equipment(
    equipment_id: int,
    target_well_id: int = Form(...),
    transfer_at: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_admin=Depends(get_current_admin),
):
    """
    Атомарний перевод обладнання на іншу свердловину.
    Закриває поточну установку і створює нову в одній транзакції.
    """
    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        raise HTTPException(status_code=404, detail="Обладнання не знайдено")

    # Перевіряємо поточну установку
    active_installation = db.query(EquipmentInstallation).filter(
        EquipmentInstallation.equipment_id == equipment_id,
        EquipmentInstallation.removed_at.is_(None),
    ).first()
    if not active_installation:
        raise HTTPException(status_code=400, detail="Обладнання не встановлено на жодній свердловині")

    if active_installation.well_id == target_well_id:
        raise HTTPException(status_code=400, detail="Обладнання вже на цій свердловині")

    # Перевіряємо цільову свердловину
    target_well = db.query(Well).filter(Well.id == target_well_id).first()
    if not target_well:
        raise HTTPException(status_code=404, detail=f"Свердловина {target_well_id} не знайдена")

    # Час переводу
    dt_transfer = datetime.now()
    if transfer_at and transfer_at.strip():
        try:
            dt_transfer = datetime.fromisoformat(transfer_at)
        except ValueError:
            pass

    # Атомарна операція: закриваємо стару + відкриваємо нову
    active_installation.removed_at = dt_transfer
    active_installation.removed_by = str(current_admin.id)
    if notes:
        active_installation.notes = (
            (active_installation.notes or "")
            + f"\nПеревод на свердловину {target_well.number}: {notes}"
        )

    new_installation = EquipmentInstallation(
        equipment_id=equipment_id,
        well_id=target_well_id,
        installed_at=dt_transfer,
        removed_at=None,
        installed_by=str(current_admin.id),
        installation_location=active_installation.installation_location,
        notes=f"Перевод зі свердловини {active_installation.well_id}. {notes or ''}".strip(),
    )
    db.add(new_installation)

    # Оновлюємо обладнання
    equipment.current_location = f"Свердловина {target_well.number}"

    db.commit()

    return JSONResponse({
        "success": True,
        "message": f"Обладнання переведено на свердловину {target_well.number}",
        "old_installation_id": active_installation.id,
        "new_installation_id": new_installation.id,
    })


# ======================================================================================
# API: Зміна статусу/стану
# ======================================================================================

@router.post("/api/equipment/{equipment_id}/update_status")
async def update_equipment_status(
    equipment_id: int,
    new_status: str = Form(...),  # available, installed, maintenance, broken, lost
    new_condition: Optional[str] = Form(None),  # working, broken, needs_repair
    new_location: Optional[str] = Form(None),  # текстова локація
    notes: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_admin = Depends(get_current_admin),
):
    """
    Оновлення статусу/стану обладнання

    Логіка локацій:
    - available → location = "Склад" (якщо не вказано інше)
    - maintenance → location = "Майстерня" / "Ремонт" / "Консервація"
    - installed → зберігається поточна локація (свердловина)
    """

    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        raise HTTPException(status_code=404, detail="Обладнання не знайдено")

    old_status = equipment.status
    old_location = equipment.current_location

    # Оновлюємо статус
    equipment.status = new_status

    # Логіка автоматичної локації
    if new_status == 'available' and not new_location:
        equipment.current_location = "Склад"
    elif new_status == 'maintenance' and not new_location:
        equipment.current_location = "Майстерня"
    elif new_location:
        equipment.current_location = new_location

    # Оновлюємо стан
    if new_condition:
        equipment.condition = new_condition

    equipment.updated_at = datetime.now()

    # Якщо переведено в статус "не на свердловині" - закриваємо активну установку
    if new_status in ['available', 'maintenance', 'broken', 'lost']:
        active_installation = db.query(EquipmentInstallation).filter(
            EquipmentInstallation.equipment_id == equipment_id,
            EquipmentInstallation.removed_at.is_(None)
        ).first()

        if active_installation:
            active_installation.removed_at = datetime.now()
            active_installation.removed_by = current_admin if isinstance(current_admin, str) else str(current_admin.id)
            if notes:
                active_installation.notes = (active_installation.notes or "") + f"\nЗміна статусу: {notes}"

    db.commit()

    return JSONResponse({
        "success": True,
        "message": f"Статус змінено: {old_status} → {new_status}. Локація: {equipment.current_location}",
        "equipment": {
            "id": equipment.id,
            "status": equipment.status,
            "condition": equipment.condition,
            "location": equipment.current_location,
        }
    })


# ======================================================================================
# API: Додати обслуговування
# ======================================================================================

@router.post("/api/equipment/{equipment_id}/add_maintenance")
async def add_maintenance(
    equipment_id: int,
    maintenance_type: str = Form(...),
    description: str = Form(...),
    cost: Optional[float] = Form(None),
    performed_by: Optional[str] = Form(None),
    maintenance_date: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_admin = Depends(get_current_admin),
):
    """
    Додати запис про обслуговування
    """

    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        raise HTTPException(status_code=404, detail="Обладнання не знайдено")

    # Парсимо дату
    if maintenance_date:
        try:
            maint_dt = datetime.fromisoformat(maintenance_date)
        except:
            maint_dt = datetime.now()
    else:
        maint_dt = datetime.now()

    # Створюємо запис
    maintenance = EquipmentMaintenance(
        equipment_id=equipment_id,
        maintenance_date=maint_dt,
        maintenance_type=maintenance_type,
        description=description,
        performed_by=performed_by or f"User {current_admin.id}",
        cost=cost,
        created_at=datetime.now(),
    )
    db.add(maintenance)

    # Оновлюємо last_maintenance_date в equipment
    equipment.last_maintenance_date = maint_dt

    db.commit()

    return JSONResponse({
        "success": True,
        "message": "Запис про обслуговування додано",
        "maintenance_id": maintenance.id,
    })


# ======================================================================================
# API: Видалення обладнання
# ======================================================================================

@router.delete("/api/equipment/{equipment_id}")
async def delete_equipment(
    equipment_id: int,
    db: Session = Depends(get_db),
    current_admin = Depends(get_current_admin),
):
    """
    Видалення обладнання (soft delete - встановлюємо deleted_at)
    """

    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        raise HTTPException(status_code=404, detail="Обладнання не знайдено")

    # Перевірка: чи не встановлено зараз
    active_installation = db.query(EquipmentInstallation).filter(
        EquipmentInstallation.equipment_id == equipment_id,
        EquipmentInstallation.removed_at.is_(None)
    ).first()

    if active_installation:
        raise HTTPException(
            status_code=400,
            detail=f"Не можна видалити обладнання яке зараз встановлено на свердловині {active_installation.well_id}. Спочатку демонтуйте."
        )

    # Soft delete
    equipment.deleted_at = datetime.now()
    db.commit()

    return JSONResponse({
        "success": True,
        "message": f"Обладнання {equipment.name} видалено",
    })

@router.get("/equipment/import", response_class=HTMLResponse)
async def equipment_import_page(
        request: Request,
):
    """Сторінка імпорту обладнання з Excel"""

    context = {
        "request": request,
    }

    return templates.TemplateResponse("equipment_import.html", context)


@router.get("/api/equipment/template")
async def download_template():
    """Завантажити Excel шаблон"""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Обладнання"

    # Заголовки
    headers = [
        "№", "Назва*", "Тип", "Серійний номер*", "Виробник",
        "Дата виробництва", "Опис", "Характеристики", "Стан", "Локація", "Примітки"
    ]

    # Стиль
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Приклад
    example = [
        1, "Електронний устьевой манометр SMOD", "Манометр", "SMOD-2024-001",
        "UNITOOL", "2024-01-15", "Система моніторингу тиску",
        "Діапазон: 0-40 МПа", "Робоче", "Склад", ""
    ]

    for col_num, value in enumerate(example, 1):
        ws.cell(row=2, column=col_num).value = value

    # Збереження в пам'ять
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Equipment_Template.xlsx"}
    )


from fastapi import UploadFile, File
from openpyxl import load_workbook
import io

@router.post("/api/equipment/import/preview")
async def preview_import(
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
):
    """
    Попередній перегляд даних з Excel
    Валідація перед імпортом
    """

    try:
        # Читання файлу
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        results = {
            "total": 0,
            "valid": 0,
            "invalid": 0,
            "duplicates": 0,
            "items": []
        }

        # Отримати існуючі серійні номери
        existing_serials = {eq.serial_number for eq in db.query(Equipment.serial_number).all()}

        # Читання даних (пропускаємо заголовок)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not any(row):  # Пустий рядок
                continue

            results["total"] += 1

            # Парсинг рядка
            item = {
                "row_num": row_num,
                "name": row[1] if len(row) > 1 else None,
                "equipment_type": row[2] if len(row) > 2 else None,
                "serial_number": row[3] if len(row) > 3 else None,
                "manufacturer": row[4] if len(row) > 4 else None,
                "manufacture_date": row[5] if len(row) > 5 else None,
                "description": row[6] if len(row) > 6 else None,
                "specifications": row[7] if len(row) > 7 else None,
                "condition": row[8] if len(row) > 8 else "working",
                "current_location": row[9] if len(row) > 9 else "Склад",
                "notes": row[10] if len(row) > 10 else None,
                "errors": [],
                "warnings": []
            }

            # Валідація
            if not item["name"]:
                item["errors"].append("Відсутня назва")

            if not item["serial_number"]:
                item["errors"].append("Відсутній серійний номер")
            elif item["serial_number"] in existing_serials:
                item["errors"].append(f"Серійний номер {item['serial_number']} вже існує")
                results["duplicates"] += 1

            # Валідація стану
            valid_conditions = ["робоче", "не робоче", "потребує ремонту", "working", "broken", "needs_repair"]
            if item["condition"] and str(item["condition"]).lower() not in valid_conditions:
                item["warnings"].append(f"Некоректний стан: {item['condition']}, буде встановлено 'working'")
                item["condition"] = "working"

            # Нормалізація стану
            condition_map = {
                "робоче": "working",
                "не робоче": "broken",
                "потребує ремонту": "needs_repair"
            }
            if item["condition"]:
                item["condition"] = condition_map.get(str(item["condition"]).lower(), item["condition"])

            # Парсинг дати
            if item["manufacture_date"]:
                if isinstance(item["manufacture_date"], date):
                    item["manufacture_date"] = item["manufacture_date"].isoformat()
                elif isinstance(item["manufacture_date"], str):
                    try:
                        parsed_date = datetime.strptime(item["manufacture_date"], "%Y-%m-%d")
                        item["manufacture_date"] = parsed_date.date().isoformat()
                    except:
                        item["warnings"].append(f"Некоректна дата: {item['manufacture_date']}")
                        item["manufacture_date"] = None

            if not item["errors"]:
                results["valid"] += 1
            else:
                results["invalid"] += 1

            results["items"].append(item)

        return JSONResponse(results)

    except Exception as e:
        return JSONResponse(
            {"error": str(e), "type": type(e).__name__},
            status_code=400
        )


@router.post("/api/equipment/import/execute")
async def execute_import(
        data: dict,
        db: Session = Depends(get_db),
        current_admin=Depends(get_current_admin),
):
    """
    Виконати імпорт обладнання
    """

    try:
        items = data.get("items", [])

        results = {
            "success": 0,
            "failed": 0,
            "skipped": 0,
            "errors": []
        }

        for item in items:
            # Пропустити якщо є помилки
            if item.get("errors"):
                results["skipped"] += 1
                continue

            try:
                # Парсинг дати
                manufacture_date = None
                if item.get("manufacture_date"):
                    try:
                        manufacture_date = datetime.fromisoformat(item["manufacture_date"]).date()
                    except:
                        pass

                # Створення обладнання
                equipment = Equipment(
                    name=item["name"],
                    equipment_type=item.get("equipment_type"),
                    serial_number=item["serial_number"],
                    manufacturer=item.get("manufacturer"),
                    manufacture_date=manufacture_date,
                    description=item.get("description"),
                    specifications=item.get("specifications"),
                    status="available",
                    condition=item.get("condition", "working"),
                    current_location=item.get("current_location", "Склад"),
                    notes=item.get("notes"),
                    created_at=datetime.now(),
                    updated_at=datetime.now(),
                )

                db.add(equipment)
                db.commit()

                results["success"] += 1

            except Exception as e:
                db.rollback()
                results["failed"] += 1
                results["errors"].append({
                    "row": item.get("row_num"),
                    "serial": item.get("serial_number"),
                    "error": str(e)
                })

        return JSONResponse(results)

    except Exception as e:
        db.rollback()
        return JSONResponse(
            {"error": str(e), "type": type(e).__name__},
            status_code=400
        )


# ======================================================================================
# API: Редактирование оборудования
# ======================================================================================

@router.post("/api/equipment/{equipment_id}/update")
async def update_equipment(
        equipment_id: int,
        name: str = Form(...),
        equipment_type: Optional[str] = Form(None),
        serial_number: Optional[str] = Form(None),
        manufacturer: Optional[str] = Form(None),
        manufacture_date: Optional[str] = Form(None),
        description: Optional[str] = Form(None),
        condition: str = Form("working"),
        notes: Optional[str] = Form(None),
        db: Session = Depends(get_db),
        current_admin=Depends(get_current_admin),
):
    """Редактирование оборудования"""

    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        raise HTTPException(status_code=404, detail="Обладнання не знайдено")

    # Обновляем поля
    equipment.name = name
    if equipment_type:
        equipment.equipment_type = equipment_type
    if serial_number:
        equipment.serial_number = serial_number
    if manufacturer:
        equipment.manufacturer = manufacturer

    if manufacture_date:
        try:
            equipment.manufacture_date = datetime.fromisoformat(manufacture_date)
        except:
            pass

    equipment.description = description
    equipment.condition = condition
    equipment.notes = notes
    equipment.updated_at = datetime.now()

    db.commit()

    return JSONResponse({
        "success": True,
        "message": "✅ Обладнання оновлено",
        "equipment_id": equipment.id
    })


# ======================================================================================
# API для работы с установками (installation)
# ======================================================================================

@router.get("/api/installation/{installation_id}")
async def get_installation_record(
        installation_id: int,
        db: Session = Depends(get_db)
):
    """Получить запись установки"""

    installation = db.query(EquipmentInstallation).filter(
        EquipmentInstallation.id == installation_id
    ).first()

    if not installation:
        raise HTTPException(status_code=404, detail="Запис встановлення не знайдено")

    return {
        "id": installation.id,
        "equipment_id": installation.equipment_id,
        "well_id": installation.well_id,
        "installed_at": installation.installed_at,
        "removed_at": installation.removed_at,
        "installation_location": installation.installation_location,
        "notes": installation.notes,
        "installed_by": installation.installed_by,
        "removed_by": installation.removed_by
    }


@router.post("/api/installation/{installation_id}/update")
async def update_installation_record(
        installation_id: int,
        installed_at: Optional[str] = Form(None),
        removed_at: Optional[str] = Form(None),
        installation_location: Optional[str] = Form(None),
        notes: Optional[str] = Form(None),
        db: Session = Depends(get_db),
        current_admin=Depends(get_current_admin)
):
    """Обновить запись установки"""

    installation = db.query(EquipmentInstallation).filter(
        EquipmentInstallation.id == installation_id
    ).first()

    if not installation:
        raise HTTPException(status_code=404, detail="Запис встановлення не знайдено")

    # Обновляем поля
    if installed_at:
        try:
            installation.installed_at = datetime.fromisoformat(installed_at.replace('T', ' '))
        except:
            pass

    if removed_at:
        try:
            installation.removed_at = datetime.fromisoformat(removed_at.replace('T', ' '))
        except:
            pass

    if installation_location:
        installation.installation_location = installation_location

    if notes:
        installation.notes = notes

    installation.updated_at = datetime.now()
    db.commit()

    return JSONResponse({
        "success": True,
        "message": "✅ Запис встановлення оновлено",
        "installation_id": installation.id
    })


@router.delete("/api/installation/{installation_id}/delete")
async def delete_installation_record(
        installation_id: int,
        db: Session = Depends(get_db),
        current_admin=Depends(get_current_admin)
):
    """Удалить запись установки (только если не активная)"""

    installation = db.query(EquipmentInstallation).filter(
        EquipmentInstallation.id == installation_id
    ).first()

    if not installation:
        raise HTTPException(status_code=404, detail="Запис встановлення не знайдено")

    # Проверка: нельзя удалять активную установку
    if installation.removed_at is None:
        raise HTTPException(
            status_code=400,
            detail="Не можна видалити активне встановлення. Спочатку демонтуйте обладнання."
        )

    db.delete(installation)
    db.commit()

    return JSONResponse({
        "success": True,
        "message": f"✅ Запис встановлення видалено",
        "installation_id": installation_id
    })