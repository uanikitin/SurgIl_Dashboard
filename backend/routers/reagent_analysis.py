"""
Анализ эффективности реагента — API и HTML страница.

Два роутера:
  - router      (prefix=/api/reagent-analysis) — JSON API
  - pages_router (без prefix)                  — HTML страница /reagent-analysis

Подключается в app.py:
    app.include_router(reagent_analysis_router)
    app.include_router(reagent_analysis_pages_router)
"""
from __future__ import annotations

import io
import logging
import time as time_module
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, Query, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field

from backend.deps import get_current_user
from backend.services.reagent_effectiveness_service import (
    analyze_reagent_effectiveness,
    get_irv_detail,
    get_overlay_data,
    ReagentAnalysisConfig,
)
from backend.services.flow_rate.data_access import list_wells_with_pressure

router = APIRouter(prefix="/api/reagent-analysis", tags=["reagent-analysis"])
pages_router = APIRouter(tags=["reagent-analysis-pages"])
templates = Jinja2Templates(directory="backend/templates")
templates.env.globals["time"] = lambda: int(time_module.time())
log = logging.getLogger(__name__)


# ──────────────────── HTML PAGE ────────────────────

@pages_router.get("/reagent-analysis", response_class=HTMLResponse)
def reagent_analysis_page(
    request: Request,
    well_id: int | None = None,
    embedded: int = 0,
    current_user: str = Depends(get_current_user),
):
    """Страница анализа эффективности реагента.

    embedded=1 — режим встраивания в iframe (без шапки/nav, см. base.html).
    well_id — пресет скважины (используется во вкладке «Реагенты» отчёта).
    """
    return templates.TemplateResponse(
        "reagent_analysis.html",
        {
            "request": request,
            "current_user": current_user,
            "is_admin": True,
            "embedded": bool(embedded),
            "preset_well_id": well_id,
        },
    )


# ──────────────────── Pydantic schemas ────────────────────

class AnalysisRequest(BaseModel):
    well_id: int
    period_start: datetime
    period_end: datetime
    grace_period_min: float = Field(default=30.0, ge=0, le=120)
    pre_window_hours: float = Field(default=2.0, ge=0.5, le=12)
    merge_window_hours: float = Field(default=4.0, ge=0.5, le=24)
    max_window_days: float = Field(default=7.0, ge=1, le=30)
    dp_effect_threshold: float = Field(default=0.3, ge=0.05, le=2.0)


# ──────────────────── API ENDPOINTS ────────────────────

@router.get("/wells")
def api_wells_list():
    """Список скважин с данными давления."""
    return list_wells_with_pressure(days=90)


@router.post("/analyze")
def api_analyze(req: AnalysisRequest):
    """
    Полный анализ эффективности реагентов для скважины.
    Возвращает ИРВ, метрики, Score, шкалу.
    """
    cfg = ReagentAnalysisConfig(
        grace_period_min=req.grace_period_min,
        pre_window_hours=req.pre_window_hours,
        merge_window_hours=req.merge_window_hours,
        max_window_days=req.max_window_days,
        dp_effect_threshold=req.dp_effect_threshold,
    )
    try:
        result = analyze_reagent_effectiveness(
            well_id=req.well_id,
            period_start=req.period_start,
            period_end=req.period_end,
            cfg=cfg,
        )
    except Exception as e:
        log.exception("reagent analysis failed: well_id=%d", req.well_id)
        raise HTTPException(status_code=500, detail=str(e))
    return result


@router.get("/irv-detail")
def api_irv_detail(
    well_id: int = Query(...),
    event_time: datetime = Query(...),
    merge_window_hours: float = Query(default=4.0, ge=0.5, le=24),
    grace_period_min: float = Query(default=30.0, ge=0, le=120),
):
    """
    Детальный анализ одного ИРВ с данными для графиков.
    """
    cfg = ReagentAnalysisConfig(
        merge_window_hours=merge_window_hours,
        grace_period_min=grace_period_min,
    )
    result = get_irv_detail(well_id, event_time, cfg)
    if result is None:
        raise HTTPException(status_code=404, detail="ИРВ не найден")
    return result


@router.post("/overlay")
def api_overlay(req: AnalysisRequest, reagent: str = Query(None)):
    """
    Overlay: ΔP кривые всех ИРВ для одного реагента.
    Ось X = часы от вброса. Для наложения на один график.
    """
    cfg = ReagentAnalysisConfig(
        grace_period_min=req.grace_period_min,
        pre_window_hours=req.pre_window_hours,
        merge_window_hours=req.merge_window_hours,
        max_window_days=req.max_window_days,
        dp_effect_threshold=req.dp_effect_threshold,
    )
    return get_overlay_data(
        well_id=req.well_id,
        period_start=req.period_start,
        period_end=req.period_end,
        reagent_name=reagent,
        cfg=cfg,
    )


@router.post("/export-excel")
def api_export_excel(req: AnalysisRequest):
    """
    Excel-экспорт результатов анализа.
    Два листа: Score по реагентам + все ИРВ.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    cfg = ReagentAnalysisConfig(
        grace_period_min=req.grace_period_min,
        pre_window_hours=req.pre_window_hours,
        merge_window_hours=req.merge_window_hours,
        max_window_days=req.max_window_days,
        dp_effect_threshold=req.dp_effect_threshold,
    )
    data = analyze_reagent_effectiveness(
        well_id=req.well_id,
        period_start=req.period_start,
        period_end=req.period_end,
        cfg=cfg,
    )

    wb = openpyxl.Workbook()

    # ── Sheet 1: Scores ──
    ws1 = wb.active
    ws1.title = "Score реагентов"
    headers1 = ["Реагент", "Вбросов", "Валидных", "Score", "Уровень",
                 "Q/шт (мед.)", "ΔP прирост (мед.)", "Длит. эффекта (мед.)", "Флаги"]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="D9E2F3")
        c.alignment = Alignment(horizontal="center")

    level_colors = {"5": "C6EFCE", "4": "D5F5D5", "3": "FFEB9C", "2": "FCD5B4", "1": "FFC7CE"}
    for i, s in enumerate(data.get("scores", []), 2):
        ws1.cell(row=i, column=1, value=s["reagent"])
        ws1.cell(row=i, column=2, value=s["irv_count"])
        ws1.cell(row=i, column=3, value=s["valid_irv_count"])
        ws1.cell(row=i, column=4, value=s["score"])
        cell_level = ws1.cell(row=i, column=5, value=s["level_name"])
        color = level_colors.get(str(s["level"]), "FFFFFF")
        cell_level.fill = PatternFill("solid", fgColor=color)
        ws1.cell(row=i, column=6, value=s.get("median_q_per_unit"))
        ws1.cell(row=i, column=7, value=s.get("median_dp_gain"))
        ws1.cell(row=i, column=8, value=s.get("median_effect_hours"))
        ws1.cell(row=i, column=9, value=", ".join(s.get("flags", [])))

    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    # ── Sheet 2: IRV details ──
    ws2 = wb.create_sheet("ИРВ")
    headers2 = ["Дата вброса", "Реагент", "Кол-во", "Объединено", "Начало ИРВ", "Конец ИРВ",
                 "Длит. (ч)", "Q_cum", "Q/шт", "ΔP прирост", "Эффект (ч)", "M5 (%)",
                 "Baseline ΔP", "Avg ΔP", "Avg Q", "Точек", "Причина исключения"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="D9E2F3")

    for i, r in enumerate(data.get("irv_results", []), 2):
        m = r["metrics"]
        ws2.cell(row=i, column=1, value=r["event_time"])
        ws2.cell(row=i, column=2, value=r["reagent"])
        ws2.cell(row=i, column=3, value=r["qty"])
        ws2.cell(row=i, column=4, value=r["merged_count"])
        ws2.cell(row=i, column=5, value=r["t_start"])
        ws2.cell(row=i, column=6, value=r["t_end"])
        ws2.cell(row=i, column=7, value=r["duration_hours"])
        ws2.cell(row=i, column=8, value=m.get("q_cumulative"))
        ws2.cell(row=i, column=9, value=m.get("q_per_unit"))
        ws2.cell(row=i, column=10, value=m.get("dp_gain"))
        ws2.cell(row=i, column=11, value=m.get("effect_duration_hours"))
        ws2.cell(row=i, column=12, value=m.get("utilisation_pct"))
        ws2.cell(row=i, column=13, value=m.get("baseline_dp"))
        ws2.cell(row=i, column=14, value=m.get("avg_dp"))
        ws2.cell(row=i, column=15, value=m.get("avg_flow_rate"))
        ws2.cell(row=i, column=16, value=m.get("data_points"))
        ws2.cell(row=i, column=17, value=m.get("invalid_reason"))

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"reagent_analysis_well{req.well_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
