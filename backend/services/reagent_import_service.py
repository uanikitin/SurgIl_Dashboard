# backend/services/reagent_import_service.py
"""
Сервис для импорта реагентов из Excel файлов.
Поддерживает:
- Импорт поступлений (reagent_supplies)
- Импорт инвентаризаций (reagent_inventory)
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from backend.models.reagent_catalog import ReagentCatalog
from backend.models.reagent_inventory import ReagentInventorySnapshot
from backend.models.reagents import ReagentSupply
from backend.services.reagent_balance_service import ReagentBalanceService


@dataclass
class ImportError:
    """Ошибка импорта для одной строки."""
    row: int
    column: str
    message: str


@dataclass
class ImportRow:
    """Валидная строка для импорта."""
    row_num: int
    date: datetime
    reagent: str
    qty: Decimal
    unit: Optional[str] = None
    comment: Optional[str] = None
    # Для инвентаризации
    calculated_qty: Optional[Decimal] = None
    discrepancy: Optional[Decimal] = None


@dataclass
class ImportResult:
    """Результат валидации/импорта."""
    success: bool = False
    total_rows: int = 0
    valid_rows: List[ImportRow] = field(default_factory=list)
    errors: List[ImportError] = field(default_factory=list)
    imported_count: int = 0
    message: str = ""


class ReagentImportService:
    """Сервис импорта реагентов из Excel."""

    # =========================================================================
    # ГЕНЕРАЦИЯ ШАБЛОНОВ
    # =========================================================================

    @staticmethod
    def generate_supply_template() -> bytes:
        """Генерирует Excel шаблон для импорта поступлений."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Поступлення"

        # Заголовки
        headers = [
            ("A", "Дата", 15, "Формат: ДД.ММ.ГГГГ або ГГГГ-ММ-ДД"),
            ("B", "Реагент", 25, "Назва реагента з каталогу"),
            ("C", "Кількість", 15, "Число > 0"),
            ("D", "Одиниця", 12, "Опціонально (за замовчуванням з каталогу)"),
            ("E", "Коментар", 40, "Опціонально"),
        ]

        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        hint_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        hint_font = Font(italic=True, color="666666", size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заполняем заголовки
        for col, (letter, title, width, hint) in enumerate(headers, 1):
            # Заголовок
            cell = ws.cell(row=1, column=col, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            # Подсказка
            hint_cell = ws.cell(row=2, column=col, value=hint)
            hint_cell.fill = hint_fill
            hint_cell.font = hint_font
            hint_cell.alignment = Alignment(horizontal="center", wrap_text=True)
            hint_cell.border = thin_border

            # Ширина колонки
            ws.column_dimensions[letter].width = width

        # Примеры данных
        examples = [
            (datetime.now(), "SMOD", 100, "шт", "Поставка від постачальника"),
            (datetime.now(), "HCI-100", 50.5, "л", ""),
            (datetime.now(), "НТФ", 25, "кг", "Терміновий заказ"),
        ]

        example_font = Font(color="808080", italic=True)
        for row_idx, (date, reagent, qty, unit, comment) in enumerate(examples, 3):
            ws.cell(row=row_idx, column=1, value=date.strftime("%d.%m.%Y")).font = example_font
            ws.cell(row=row_idx, column=2, value=reagent).font = example_font
            ws.cell(row=row_idx, column=3, value=qty).font = example_font
            ws.cell(row=row_idx, column=4, value=unit).font = example_font
            ws.cell(row=row_idx, column=5, value=comment).font = example_font

        # Инструкция на втором листе
        ws_help = wb.create_sheet("Інструкція")
        instructions = [
            "ІНСТРУКЦІЯ ПО ЗАПОВНЕННЮ",
            "",
            "1. Заповнюйте дані починаючи з 3-го рядка (рядки 1-2 - заголовки)",
            "2. Дата: формат ДД.ММ.ГГГГ (наприклад: 15.01.2026) або ГГГГ-ММ-ДД",
            "3. Реагент: точна назва з каталогу (регістр має значення)",
            "4. Кількість: число більше 0, можна з десятковими (через . або ,)",
            "5. Одиниця: якщо не вказано - береться з каталогу реагентів",
            "6. Коментар: довільний текст (необов'язково)",
            "",
            "ВАЖЛИВО:",
            "- Не змінюйте структуру файлу (не додавайте/видаляйте колонки)",
            "- Видаліть приклади перед завантаженням",
            "- Перевірте назви реагентів перед імпортом",
        ]
        for idx, line in enumerate(instructions, 1):
            cell = ws_help.cell(row=idx, column=1, value=line)
            if idx == 1:
                cell.font = Font(bold=True, size=14)
            elif line.startswith("ВАЖЛИВО"):
                cell.font = Font(bold=True, color="FF0000")
        ws_help.column_dimensions["A"].width = 80

        # Сохраняем в байты
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_inventory_template() -> bytes:
        """Генерирует Excel шаблон для импорта инвентаризаций."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Інвентаризація"

        # Заголовки
        headers = [
            ("A", "Дата", 15, "Формат: ДД.ММ.ГГГГ або ГГГГ-ММ-ДД"),
            ("B", "Реагент", 25, "Назва реагента з каталогу"),
            ("C", "Фактична кількість", 18, "Число >= 0"),
            ("D", "Коментар", 40, "Опціонально"),
        ]

        # Стили
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        hint_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        hint_font = Font(italic=True, color="666666", size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заполняем заголовки
        for col, (letter, title, width, hint) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            hint_cell = ws.cell(row=2, column=col, value=hint)
            hint_cell.fill = hint_fill
            hint_cell.font = hint_font
            hint_cell.alignment = Alignment(horizontal="center", wrap_text=True)
            hint_cell.border = thin_border

            ws.column_dimensions[letter].width = width

        # Примеры
        examples = [
            (datetime.now(), "SMOD", 85, "Перерахунок на складі"),
            (datetime.now(), "HCI-100", 42.5, ""),
            (datetime.now(), "НТФ", 0, "Залишків немає"),
        ]

        example_font = Font(color="808080", italic=True)
        for row_idx, (date, reagent, qty, comment) in enumerate(examples, 3):
            ws.cell(row=row_idx, column=1, value=date.strftime("%d.%m.%Y")).font = example_font
            ws.cell(row=row_idx, column=2, value=reagent).font = example_font
            ws.cell(row=row_idx, column=3, value=qty).font = example_font
            ws.cell(row=row_idx, column=4, value=comment).font = example_font

        # Инструкция
        ws_help = wb.create_sheet("Інструкція")
        instructions = [
            "ІНСТРУКЦІЯ ПО ЗАПОВНЕННЮ ІНВЕНТАРИЗАЦІЇ",
            "",
            "1. Заповнюйте дані починаючи з 3-го рядка",
            "2. Дата: дата проведення інвентаризації",
            "3. Реагент: точна назва з каталогу",
            "4. Фактична кількість: реальний залишок при перерахунку (>= 0)",
            "5. Коментар: примітки (необов'язково)",
            "",
            "ЯК ПРАЦЮЄ ІНВЕНТАРИЗАЦІЯ:",
            "- Система автоматично розрахує очікуваний залишок на дату",
            "- Різниця (факт - розрахунок) буде записана як розходження",
            "- Позитивне розходження = надлишок",
            "- Негативне розходження = нестача",
            "",
            "ВАЖЛИВО:",
            "- Можна імпортувати за минулі дати",
            "- Інвентаризація скидає базу розрахунку для подальших операцій",
        ]
        for idx, line in enumerate(instructions, 1):
            cell = ws_help.cell(row=idx, column=1, value=line)
            if idx == 1:
                cell.font = Font(bold=True, size=14)
            elif line.startswith("ВАЖЛИВО") or line.startswith("ЯК ПРАЦЮЄ"):
                cell.font = Font(bold=True, color="FF0000" if "ВАЖЛИВО" in line else "0070C0")
        ws_help.column_dimensions["A"].width = 80

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # =========================================================================
    # ПАРСИНГ И ВАЛИДАЦИЯ
    # =========================================================================

    @staticmethod
    def _parse_date(value) -> Optional[datetime]:
        """Парсит дату из различных форматов."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, str):
            value = value.strip()
            formats = ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"]
            for fmt in formats:
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue
        return None

    @staticmethod
    def _parse_decimal(value) -> Optional[Decimal]:
        """Парсит число в Decimal."""
        if value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return Decimal(str(value))
            if isinstance(value, str):
                # Заменяем запятую на точку
                value = value.strip().replace(",", ".")
                return Decimal(value)
            if isinstance(value, Decimal):
                return value
        except (InvalidOperation, ValueError):
            pass
        return None

    @classmethod
    def validate_supply_file(
        cls,
        db: Session,
        file_content: bytes,
    ) -> ImportResult:
        """
        Валидирует Excel файл с поступлениями.
        Возвращает результат с валидными строками и ошибками.
        """
        result = ImportResult()

        try:
            wb = load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            result.message = f"Помилка читання файлу: {str(e)}"
            return result

        # Получаем список реагентов из каталога
        catalog = {r.name: r for r in db.query(ReagentCatalog).filter(ReagentCatalog.is_active == True).all()}

        # Читаем строки (пропускаем 2 строки заголовков)
        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            # Пропускаем пустые строки
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            result.total_rows += 1

            # Парсим поля
            date_val = row[0] if len(row) > 0 else None
            reagent_val = row[1] if len(row) > 1 else None
            qty_val = row[2] if len(row) > 2 else None
            unit_val = row[3] if len(row) > 3 else None
            comment_val = row[4] if len(row) > 4 else None

            row_errors = []

            # Валидация даты
            parsed_date = cls._parse_date(date_val)
            if not parsed_date:
                row_errors.append(ImportError(row_idx, "A", f"Невірний формат дати: {date_val}"))

            # Валидация реагента
            reagent_name = str(reagent_val).strip() if reagent_val else None
            if not reagent_name:
                row_errors.append(ImportError(row_idx, "B", "Реагент не вказано"))
            elif reagent_name not in catalog:
                row_errors.append(ImportError(row_idx, "B", f"Реагент '{reagent_name}' не знайдено в каталозі"))

            # Валидация количества
            parsed_qty = cls._parse_decimal(qty_val)
            if parsed_qty is None:
                row_errors.append(ImportError(row_idx, "C", f"Невірний формат кількості: {qty_val}"))
            elif parsed_qty <= 0:
                row_errors.append(ImportError(row_idx, "C", f"Кількість повинна бути > 0: {qty_val}"))

            # Единица измерения
            unit = None
            if unit_val:
                unit = str(unit_val).strip()
            elif reagent_name and reagent_name in catalog:
                unit = catalog[reagent_name].default_unit

            # Комментарий
            comment = str(comment_val).strip() if comment_val else None

            if row_errors:
                result.errors.extend(row_errors)
            else:
                result.valid_rows.append(ImportRow(
                    row_num=row_idx,
                    date=parsed_date,
                    reagent=reagent_name,
                    qty=parsed_qty,
                    unit=unit,
                    comment=comment,
                ))

        result.success = len(result.errors) == 0 and len(result.valid_rows) > 0
        if result.success:
            result.message = f"Валідація пройдена: {len(result.valid_rows)} рядків готові до імпорту"
        elif len(result.valid_rows) > 0:
            result.message = f"Знайдено {len(result.errors)} помилок. {len(result.valid_rows)} рядків валідні."
        else:
            result.message = f"Імпорт неможливий: {len(result.errors)} помилок"

        return result

    @classmethod
    def validate_inventory_file(
        cls,
        db: Session,
        file_content: bytes,
    ) -> ImportResult:
        """
        Валидирует Excel файл с инвентаризациями.
        Рассчитывает expected_qty и discrepancy для каждой строки.
        """
        result = ImportResult()

        try:
            wb = load_workbook(BytesIO(file_content), data_only=True)
            ws = wb.active
        except Exception as e:
            result.message = f"Помилка читання файлу: {str(e)}"
            return result

        catalog = {r.name: r for r in db.query(ReagentCatalog).filter(ReagentCatalog.is_active == True).all()}

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            result.total_rows += 1

            date_val = row[0] if len(row) > 0 else None
            reagent_val = row[1] if len(row) > 1 else None
            qty_val = row[2] if len(row) > 2 else None
            comment_val = row[3] if len(row) > 3 else None

            row_errors = []

            # Валидация даты
            parsed_date = cls._parse_date(date_val)
            if not parsed_date:
                row_errors.append(ImportError(row_idx, "A", f"Невірний формат дати: {date_val}"))

            # Валидация реагента
            reagent_name = str(reagent_val).strip() if reagent_val else None
            if not reagent_name:
                row_errors.append(ImportError(row_idx, "B", "Реагент не вказано"))
            elif reagent_name not in catalog:
                row_errors.append(ImportError(row_idx, "B", f"Реагент '{reagent_name}' не знайдено в каталозі"))

            # Валидация количества (для инвентаризации может быть 0)
            parsed_qty = cls._parse_decimal(qty_val)
            if parsed_qty is None:
                row_errors.append(ImportError(row_idx, "C", f"Невірний формат кількості: {qty_val}"))
            elif parsed_qty < 0:
                row_errors.append(ImportError(row_idx, "C", f"Кількість не може бути від'ємною: {qty_val}"))

            comment = str(comment_val).strip() if comment_val else None

            if row_errors:
                result.errors.extend(row_errors)
            else:
                # Рассчитываем ожидаемый остаток
                calculated_qty = None
                discrepancy = None

                if parsed_date and reagent_name:
                    balance_data = ReagentBalanceService.get_current_balance(
                        db, reagent_name, parsed_date
                    )
                    calculated_qty = Decimal(str(balance_data.get("current_balance", 0)))
                    discrepancy = parsed_qty - calculated_qty

                result.valid_rows.append(ImportRow(
                    row_num=row_idx,
                    date=parsed_date,
                    reagent=reagent_name,
                    qty=parsed_qty,
                    unit=catalog[reagent_name].default_unit if reagent_name in catalog else "шт",
                    comment=comment,
                    calculated_qty=calculated_qty,
                    discrepancy=discrepancy,
                ))

        result.success = len(result.errors) == 0 and len(result.valid_rows) > 0
        if result.success:
            result.message = f"Валідація пройдена: {len(result.valid_rows)} рядків готові до імпорту"
        elif len(result.valid_rows) > 0:
            result.message = f"Знайдено {len(result.errors)} помилок. {len(result.valid_rows)} рядків валідні."
        else:
            result.message = f"Імпорт неможливий: {len(result.errors)} помилок"

        return result

    # =========================================================================
    # ИМПОРТ В БАЗУ
    # =========================================================================

    @classmethod
    def import_supplies(
        cls,
        db: Session,
        file_content: bytes,
        created_by: Optional[str] = None,
    ) -> ImportResult:
        """
        Импортирует поступления из Excel в базу данных.
        """
        # Сначала валидируем
        result = cls.validate_supply_file(db, file_content)

        if not result.valid_rows:
            return result

        # Получаем каталог для связи
        catalog = {r.name: r for r in db.query(ReagentCatalog).filter(ReagentCatalog.is_active == True).all()}

        imported = 0
        for row in result.valid_rows:
            try:
                supply = ReagentSupply(
                    reagent=row.reagent,
                    reagent_id=catalog[row.reagent].id if row.reagent in catalog else None,
                    qty=row.qty,
                    unit=row.unit or "шт",
                    received_at=row.date,
                    comment=row.comment,
                    source=f"Excel import by {created_by}" if created_by else "Excel import",
                )
                db.add(supply)
                imported += 1
            except Exception as e:
                result.errors.append(ImportError(row.row_num, "-", f"Помилка збереження: {str(e)}"))

        try:
            db.commit()
            result.imported_count = imported
            result.success = imported > 0
            result.message = f"Успішно імпортовано {imported} записів поступлень"
        except Exception as e:
            db.rollback()
            result.success = False
            result.message = f"Помилка збереження в базу: {str(e)}"

        return result

    @classmethod
    def import_inventory(
        cls,
        db: Session,
        file_content: bytes,
        created_by: Optional[str] = None,
    ) -> ImportResult:
        """
        Импортирует инвентаризации из Excel в базу данных.
        """
        result = cls.validate_inventory_file(db, file_content)

        if not result.valid_rows:
            return result

        catalog = {r.name: r for r in db.query(ReagentCatalog).filter(ReagentCatalog.is_active == True).all()}

        imported = 0
        for row in result.valid_rows:
            try:
                inventory = ReagentInventorySnapshot(
                    reagent=row.reagent,
                    reagent_id=catalog[row.reagent].id if row.reagent in catalog else None,
                    qty=row.qty,
                    calculated_qty=row.calculated_qty,
                    discrepancy=row.discrepancy,
                    unit=row.unit or "шт",
                    snapshot_at=row.date,
                    comment=row.comment,
                    created_by=created_by or "Excel import",
                )
                db.add(inventory)
                imported += 1
            except Exception as e:
                result.errors.append(ImportError(row.row_num, "-", f"Помилка збереження: {str(e)}"))

        try:
            db.commit()
            result.imported_count = imported
            result.success = imported > 0
            result.message = f"Успішно імпортовано {imported} записів інвентаризації"
        except Exception as e:
            db.rollback()
            result.success = False
            result.message = f"Помилка збереження в базу: {str(e)}"

        return result
