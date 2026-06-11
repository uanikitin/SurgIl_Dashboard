"""
ocr_day_report_pdf.py
=====================

Распознавание СКАНА суточной сводки по скважинам (PDF от Canon iR2520, без
текстового слоя) → канонический .xlsx в формате, который ест существующий
parsing_day_report_UZKOR.py.

Гибридная схема (подтверждена на 01.08.2025, end-to-end ≈ 90% по ячейкам,
ключевые поля 100%):

    геометрия (cv2)            : поворот 90° → авто-дескью → детекция сетки
    Tesseract (локально)       : давления (устье/затруб/шлейф/статика), простой
    vision-LLM (Anthropic)     : номер скважины, штуцер, дебит  ← курсивный шрифт
    справочник well→ГСП        : ГСП не распознаём (стабилен на 100% по истории)

Открытия по формату PDF (НЕ как в Excel!):
  * блок = 9 строк данных + 1 пустая строка-разделитель (10 интервалов);
  * дебит газа ОДИН (= q_work, «с учётом раб.времени»); Excel-овский q_total —
    обратный пересчёт через коэф.расходомера, на бумаге его нет;
  * дебит и штуцер печатаются ЦЕЛЫМИ; «0» штуцера/простоя = пустая ячейка;
  * № ГСП печатается по центру под-блока (позиционно не читаем) → берём из
    справочника well→ГСП.

Зависимости: pdftoppm (poppler), tesseract (+rus/eng), opencv-python, pytesseract,
openpyxl; для vision — пакет `anthropic` и переменная окружения ANTHROPIC_API_KEY.

CLI:
    python -m backend.utils.ocr_day_report_pdf <wells.pdf> [-o out.xlsx] [--no-vision]
"""
from __future__ import annotations

import argparse
import base64
import json
import logging
import os
import re
import subprocess
import sys
import tempfile
from datetime import date
from pathlib import Path
from typing import Optional

import cv2
import numpy as np
import openpyxl
import pytesseract

log = logging.getLogger("ocr_day_report_pdf")

# ───────────────────────────── справочники ──────────────────────────────
_DATA = Path(__file__).parent / "ocr_data"
_WELL2GGU: dict[str, str] = json.loads((_DATA / "well2ggu.json").read_text("utf-8"))["stable"]
_REGISTRY: set[str] = set(json.loads((_DATA / "well_registry.json").read_text("utf-8")))
_REG_NUM = sorted(int(x) for x in _REGISTRY if x.isdigit())
_CHOKES = {0, 6, 7, 8, 9, 10, 11, 12, 13, 14, 16}

# Геометрия блока: 9 строк данных + пустой разделитель (10 интервалов)
_ROWS = ["ggu", "well", "choke", "p_wh", "p_an", "p_fl", "q_gas", "shut", "p_st", "_blank"]
_PMAX = {"p_wh": 200, "p_an": 200, "p_fl": 60, "p_st": 250}
# openpyxl-строки блоков в каноническом xlsx (как ждёт парсер)
_HEAD = [3, 14, 25, 36]
_LABELS = {
    0: "№ ГСП / GGU No.", 1: "Номера скважин / Well No.", 2: "Диаметр штуцера",
    3: "Давление на устье", 4: "Затрубное давление", 5: "Давление в шлейфе",
    6: "Дебит газа total", 7: "Дебит газа", 8: "Время простоя", 9: "Статическое давление",
}
_VISION_FIELDS = {"well", "choke", "q_gas"}   # что отдаём vision

# ───────────────────────────── геометрия ────────────────────────────────
def _render(pdf: str, dpi: int = 400) -> np.ndarray:
    d = tempfile.mkdtemp()
    base = os.path.join(d, "p")
    subprocess.run(["pdftoppm", "-png", "-r", str(dpi), pdf, base], check=True)
    return cv2.imread(base + "-1.png", cv2.IMREAD_GRAYSCALE)


def _deskew(img: np.ndarray) -> np.ndarray:
    _, bw = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    h, w = img.shape
    horiz = cv2.morphologyEx(bw, cv2.MORPH_OPEN,
                             cv2.getStructuringElement(cv2.MORPH_RECT, (max(60, w // 40), 1)))
    lines = cv2.HoughLinesP(horiz, 1, np.pi / 720, threshold=200,
                            minLineLength=w // 6, maxLineGap=20)
    angs = []
    if lines is not None:
        for x1, y1, x2, y2 in lines[:, 0]:
            a = np.degrees(np.arctan2(y2 - y1, x2 - x1))
            if abs(a) < 10:
                angs.append(a)
    ang = float(np.median(angs)) if angs else 0.0
    M = cv2.getRotationMatrix2D((w / 2, h / 2), ang, 1.0)
    return cv2.warpAffine(img, M, (w, h), flags=cv2.INTER_CUBIC, borderValue=255)


def _upright(img: np.ndarray) -> np.ndarray:
    """Скан повёрнут — ставим таблицу горизонтально, затем выпрямляем."""
    return _deskew(cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE))


def _group(idx, gap=8):
    if len(idx) == 0:
        return []
    g = [[idx[0]]]
    for v in idx[1:]:
        g[-1].append(v) if v - g[-1][-1] <= gap else g.append([v])
    return [int(np.mean(x)) for x in g]


def _hlines(img):
    _, bw = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    w = img.shape[1]
    horiz = cv2.morphologyEx(bw, cv2.MORPH_OPEN,
                             cv2.getStructuringElement(cv2.MORPH_RECT, (150, 1)))
    return bw, _group(np.where((horiz > 0).sum(axis=1) > 0.18 * w)[0])


def _vlines(bw, y0, y1):
    band = cv2.morphologyEx(bw[y0:y1, :], cv2.MORPH_OPEN,
                            cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40)))
    return _group(np.where((band > 0).sum(axis=0) > 0.5 * (y1 - y0))[0])


def _block_vlines(bw, yb, width):
    """Вертикальные линии блока без колонки-подписи (отбрасываем по последнему
    крупному зазору в левой части)."""
    vl = _vlines(bw, yb[0], yb[10])
    big = [i for i in range(len(vl) - 1) if vl[i + 1] - vl[i] > 180 and vl[i] < 0.45 * width]
    ds = (big[-1] + 1) if big else 0
    return vl[ds:]


# ───────────────────────────── Tesseract ────────────────────────────────
def _ocr_num(crop) -> str:
    if crop.size == 0:
        return ""
    c = cv2.resize(crop, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)
    c = cv2.copyMakeBorder(c, 15, 15, 15, 15, cv2.BORDER_CONSTANT, value=255)
    _, c = cv2.threshold(c, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    for psm in (7, 8, 13):
        t = pytesseract.image_to_string(
            c, config=f"--psm {psm} -c tessedit_char_whitelist=0123456789,.").strip()
        if t:
            return t.replace(" ", "").replace(",", ".").strip(".")
    return ""


def _ocr_annular(crop) -> str:
    if crop.size == 0:
        return ""
    c = cv2.resize(crop, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)
    c = cv2.copyMakeBorder(c, 15, 15, 15, 15, cv2.BORDER_CONSTANT, value=255)
    t = pytesseract.image_to_string(c, lang="rus", config="--psm 7").strip().lower()
    if "па" in t:
        return "пак"
    return _ocr_num(crop)


# ───────────────────────────── vision (Anthropic) ───────────────────────
def _strip_png(img, ry0, ry1, vl) -> bytes:
    """Полоса строки с красными разделителями ячеек → PNG-байты."""
    x0, x1 = vl[0] - 2, vl[-1] + 2
    strip = cv2.cvtColor(img[ry0:ry1, x0:x1], cv2.COLOR_GRAY2BGR)
    for v in vl:
        cv2.line(strip, (v - x0, 0), (v - x0, strip.shape[0]), (0, 0, 255), 1)
    strip = cv2.resize(strip, None, fx=2.2, fy=2.2, interpolation=cv2.INTER_CUBIC)
    return cv2.imencode(".png", strip)[1].tobytes()


_VISION_PROMPT = (
    "На изображении — ОДНА строка таблицы, разделённая на {n} ячеек красными "
    "вертикальными линиями, читать слева направо. Верни СТРОГО JSON-массив из "
    "ровно {n} строковых элементов — содержимое каждой ячейки по порядку. "
    "Пустая ячейка → \"\". {hint} "
    "Не угадывай и не вычисляй — верни только то, что напечатано. "
    "Ответ — только JSON-массив, без пояснений."
)
_HINTS = {
    "well": "Значения — номера скважин (целые) либо токены '12тр','12зтр','NA1','NA2'.",
    "choke": "Значения — диаметр штуцера, малые целые (6..16).",
    "q_gas": "Значения — дебит газа, целые числа.",
}


def _vision_read(client, model, png: bytes, field: str, n: int) -> list[str]:
    b64 = base64.standard_b64encode(png).decode()
    msg = client.messages.create(
        model=model, max_tokens=1024,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": b64}},
            {"type": "text", "text": _VISION_PROMPT.format(n=n, hint=_HINTS[field])},
        ]}],
    )
    txt = "".join(b.text for b in msg.content if getattr(b, "type", "") == "text")
    m = re.search(r"\[.*\]", txt, re.S)
    arr = json.loads(m.group(0)) if m else []
    arr = [("" if v is None else str(v).strip()) for v in arr]
    if len(arr) != n:                       # выравниваем длину под число ячеек
        arr = (arr + [""] * n)[:n]
    return arr


def _get_vision_client():
    if not os.getenv("ANTHROPIC_API_KEY"):
        return None, None
    try:
        import anthropic
    except ImportError:
        log.warning("пакет anthropic не установлен — vision недоступен")
        return None, None
    model = os.getenv("OCR_VISION_MODEL", "claude-3-5-sonnet-latest")
    return anthropic.Anthropic(), model


# ───────────────────────── доменная коррекция ───────────────────────────
def _to_f(s) -> Optional[float]:
    if s in (None, ""):
        return None
    try:
        return float(str(s).replace(",", "."))
    except ValueError:
        return None


def _dist(a, b):
    return sum(1 for x, y in zip(a, b) if x != y)


def _snap_well(s):
    if not s:
        return None
    if s in _REGISTRY:
        return s
    d = re.sub(r"\D", "", s)
    if not d:
        return s
    if d in _REGISTRY:
        return d
    same = [str(n) for n in _REG_NUM if len(str(n)) == len(d)]
    if same:
        best = min(same, key=lambda n: _dist(n, d))
        if _dist(best, d) <= 1:
            return best
    return d


def _snap_choke(s):
    f = _to_f(s)
    if f is None:
        return None
    i = int(round(f))
    if i in _CHOKES:
        return i
    cand = [c for c in _CHOKES if len(str(c)) == len(str(i))]
    if cand:
        best = min(cand, key=lambda c: _dist(str(c), str(i)))
        if _dist(str(best), str(i)) <= 1:
            return best
    n = min(_CHOKES, key=lambda c: abs(c - i))
    return n if abs(n - i) <= 1 else i


def _fix_press(s, vmax):
    f = _to_f(s)
    if f is None:
        return None
    if f <= vmax:
        return round(f, 1)
    d = re.sub(r"\D", "", str(s))
    if len(d) >= 2:
        for div in (10.0, 100.0):
            c = round(int(d) / div, 1)
            if c <= vmax:
                return c
    return round(f, 1)


def _ggu_marker(well):
    g = _WELL2GGU.get(str(well))
    if not g:
        return None
    if g == "БВН":
        return "БВН"
    m = re.match(r"ГСП-(\d+)", g)
    return int(m.group(1)) if m else g


# ───────────────────────────── извлечение ───────────────────────────────
def extract_blocks(pdf: str, use_vision: bool = True) -> list[dict]:
    """Распознать 4 блока скважинного PDF. Возвращает список словарей с сырыми
    значениями полей (списки по ячейкам)."""
    img = _upright(_render(pdf))
    bw, hl = _hlines(img)
    if len(hl) < 41:
        raise RuntimeError(f"не удалось построить сетку: найдено {len(hl)} гор. линий (нужно ≥41)")
    y = hl[:41]
    W = img.shape[1]

    client, model = (_get_vision_client() if use_vision else (None, None))
    if use_vision and client is None:
        log.warning("vision отключён (нет ключа/SDK) — well/choke/q_gas будут читаться Tesseract")

    blocks = []
    for b in range(4):
        yb = y[b * 10: b * 10 + 11]
        vl = _block_vlines(bw, yb, W)
        ncell = len(vl) - 1
        block = {"ncell": ncell}
        # vision-строки одним вызовом каждая
        vision_vals = {}
        if client is not None:
            for f in ("well", "choke", "q_gas"):
                off = _ROWS.index(f)
                png = _strip_png(img, yb[off] - 4, yb[off + 1] + 4, vl)
                try:
                    vision_vals[f] = _vision_read(client, model, png, f, ncell)
                except Exception as e:        # noqa: BLE001
                    log.warning("vision %s блок %d: %s — fallback Tesseract", f, b + 1, e)
        # построчно
        for ri, name in enumerate(_ROWS):
            if name in ("ggu", "_blank"):
                continue
            if name in vision_vals:
                block[name] = vision_vals[name]
                continue
            ry0, ry1 = yb[ri] + 3, yb[ri + 1] - 3
            ocr = _ocr_annular if name == "p_an" else _ocr_num
            block[name] = [ocr(img[ry0:ry1, vl[ci] + 3: vl[ci + 1] - 3]) for ci in range(ncell)]
        blocks.append(block)
    return blocks


# ───────────────────────────── сборка xlsx ──────────────────────────────
def write_day_sheet(ws, blocks: list[dict], sheet_date: date) -> list:
    """Записать блоки одного дня в готовый лист ws. Возвращает список спорных ячеек."""
    ws["A1"] = f"Daily information (OCR) {sheet_date.isoformat()}"
    flags = []  # (date, well, field, value) — для ручной проверки

    for b, block in enumerate(blocks):
        hr = _HEAD[b]
        for off, lab in _LABELS.items():
            ws.cell(row=hr + off, column=1, value=lab)
        wells = block.get("well", [])
        prev = None
        for i, raw_well in enumerate(wells):
            well = _snap_well(raw_well)
            if not well:
                continue
            col = 2 + i
            ws.cell(row=hr + 1, column=col, value=str(well))
            if well not in _REGISTRY:
                flags.append((sheet_date.isoformat(), well, "well", raw_well))
            g = _ggu_marker(well)
            if g is None:
                flags.append((sheet_date.isoformat(), well, "ggu", "нет в справочнике"))
            if g != prev:
                ws.cell(row=hr, column=col, value=g)
                prev = g

            def cell(key, idx=i):
                v = block.get(key, [])
                return v[idx] if idx < len(v) else None

            ws.cell(row=hr + 2, column=col, value=_snap_choke(cell("choke")))
            ws.cell(row=hr + 3, column=col, value=_fix_press(cell("p_wh"), _PMAX["p_wh"]))
            pa = cell("p_an")
            ws.cell(row=hr + 4, column=col,
                    value=("пак" if str(pa).strip().lower() == "пак" else _fix_press(pa, _PMAX["p_an"])))
            ws.cell(row=hr + 5, column=col, value=_fix_press(cell("p_fl"), _PMAX["p_fl"]))
            # бумага = q_work (целое) → строка q_work (offset 7); q_total (6) пуст
            qg = _to_f(cell("q_gas"))
            ws.cell(row=hr + 7, column=col, value=None if qg is None else int(round(qg)))
            sh = _to_f(cell("shut"))
            sh = 0 if sh is None else (1440 if sh > 1400 else sh)
            ws.cell(row=hr + 8, column=col, value=sh)
            ws.cell(row=hr + 9, column=col, value=_fix_press(cell("p_st"), _PMAX["p_st"]))

    return flags


def build_xlsx(blocks: list[dict], sheet_date: date, out_path: str) -> dict:
    """Собрать одностраничный канонический xlsx (один день)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_date.strftime("%d.%m.%y")
    flags = write_day_sheet(ws, blocks, sheet_date)
    wb.save(out_path)
    return {"out": out_path, "flags": flags}


_DATE_RX = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")


def _date_from_name(pdf: str) -> Optional[date]:
    m = _DATE_RX.search(Path(pdf).name)
    if not m:
        return None
    d, mo, y = (int(x) for x in m.groups())
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def pdf_to_xlsx(pdf: str, out: Optional[str] = None, use_vision: bool = True) -> dict:
    sheet_date = _date_from_name(pdf) or date.today()
    out = out or str(Path(pdf).with_suffix(".xlsx"))
    blocks = extract_blocks(pdf, use_vision=use_vision)
    return build_xlsx(blocks, sheet_date, out)


# ───────────────────────────────── CLI ──────────────────────────────────
def main(argv=None) -> int:
    p = argparse.ArgumentParser(description="OCR скан-PDF суточной сводки скважин → canonical .xlsx")
    p.add_argument("pdf", help="путь к PDF (скважины)")
    p.add_argument("-o", "--output", default=None)
    p.add_argument("--no-vision", action="store_true", help="без vision (всё Tesseract)")
    p.add_argument("-v", "--verbose", action="store_true")
    a = p.parse_args(argv)
    logging.basicConfig(level=logging.DEBUG if a.verbose else logging.INFO,
                        format="%(levelname)s: %(message)s")
    res = pdf_to_xlsx(a.pdf, a.output, use_vision=not a.no_vision)
    log.info("готово → %s", res["out"])
    if res["flags"]:
        log.warning("спорных ячеек на проверку: %d", len(res["flags"]))
        for f in res["flags"][:30]:
            log.warning("  %s", f)
    return 0


if __name__ == "__main__":
    sys.exit(main())
