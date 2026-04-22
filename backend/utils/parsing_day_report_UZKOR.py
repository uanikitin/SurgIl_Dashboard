"""
parse_daily_wells.py
====================

Парсер суточных сводок по газодобывающим скважинам месторождения Сургил
(формат Excel-файла «Суточная сводка <МЕСЯЦ ГОД>г.xlsx»).

Универсальный: не привязан к конкретному месяцу. Ожидает типовую
структуру листа:

    Row 1 : заголовок «Daily information ...»
    Row 3 : № ГСП / GGU No.        (блок 1: ГСП-2 + БВН)
    Row 4 : Номера скважин
    Row 5 : Диаметр штуцера, мм
    Row 6 : Давление на устье
    Row 7 : Затрубное давление
    Row 8 : Давление в шлейфе
    Row 9 : Дебит газа (общий)
    Row 10: Дебит газа (с учётом времени работы / простоев)
    Row 11: Время простоя, мин
    Row 12: Статическое давление
    Row 14-23 : Блок 2 (ГСП-1 + ГСП-4)
    Row 25-34 : Блок 3 (ГСП-3)
    Row 36-45 : Блок 4 (ГСП-5 + ГСП-6)

В строке № ГСП могут появляться несколько меток (напр. «2», «БВН» или
«1», «4») — они обозначают начало под-блока. Парсер делает forward-fill:
каждая скважина получает ГСП от ближайшей слева метки.

Каждая строка итогового CSV = одна скважина × один день.

Использование
-------------
    python parse_daily_wells.py <input.xlsx> [-o output.csv]

Пример:
    python parse_daily_wells.py "Суточная сводка МАРТ 2026г.xlsx"
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

log = logging.getLogger("parse_daily_wells")

# ────────────────────────── Конфигурация листа ──────────────────────────

# Смещения строк внутри каждого блока (относительно строки «№ ГСП»)
ROW_OFFSETS = {
    "ggu":             0,   # № ГСП / GGU No.
    "well":            1,   # Номера скважин
    "choke_mm":        2,   # Диаметр штуцера
    "p_wellhead":      3,   # Давление на устье
    "p_annular":       4,   # Затрубное давление
    "p_flowline":      5,   # Давление в шлейфе
    "q_gas_total":     6,   # Дебит газа, общий
    "q_gas_working":   7,   # Дебит газа, с учётом раб. времени
    "shutdown_min":    8,   # Время простоя, мин
    "p_static":        9,   # Статическое давление
}

# Строки, где начинаются блоки (строка «№ ГСП» для каждого из 4 блоков)
BLOCK_HEADER_ROWS = [3, 14, 25, 36]

# Диапазон колонок, в которых могут быть скважины
MIN_COL = 2
MAX_COL = 46

# Маркеры «ошибок» Excel, которые надо превращать в NaN
EXCEL_ERRORS = {"#VALUE!", "#REF!", "#DIV/0!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}

# «пак» в затрубном давлении = установлен пакер (численного значения нет)
PACKER_TOKENS = {"пак", "пакер", "pak", "packer"}


# ──────────────────────────── Утилиты ───────────────────────────────────

def _to_float(value) -> Optional[float]:
    """Преобразовать ячейку к float. Вернёт None, если невозможно."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s in EXCEL_ERRORS:
        return None
    # Запятая → точка (на случай локали)
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _well_id(value) -> Optional[str]:
    """Преобразовать номер скважины к строке. None для пустых ячеек."""
    if value is None:
        return None
    if isinstance(value, float) and value != value:  # NaN
        return None
    s = str(value).strip()
    if not s:
        return None
    # «8.0» → «8»
    try:
        f = float(s.replace(",", "."))
        if f.is_integer():
            return str(int(f))
    except ValueError:
        pass
    return s


def _is_packer(value) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in PACKER_TOKENS


# ─────────────────────────── Разбор даты ────────────────────────────────

_SHEET_DATE_RX = re.compile(r"(\d{1,2})[.,\-/](\d{1,2})[.,\-/](\d{2,4})")


def parse_sheet_date(sheet_name: str) -> Optional[date]:
    """Распарсить дату из имени листа (допускает разделители . , - /, пробелы)."""
    s = sheet_name.strip()
    m = _SHEET_DATE_RX.search(s)
    if not m:
        return None
    d, mo, y = (int(x) for x in m.groups())
    if y < 100:
        y += 2000
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def extract_date(ws) -> Optional[date]:
    """
    Получить дату листа.
    Стратегия: 1) искать datetime в первой строке рядом с подписью «Дата / Date»;
               2) fallback — распарсить имя листа.
    """
    # 1) Проходим первую строку, ищем datetime
    for cell in ws[1]:
        v = cell.value
        if isinstance(v, (datetime, date)) and not isinstance(v, bool):
            return v.date() if isinstance(v, datetime) else v
    # 2) Fallback — имя листа
    return parse_sheet_date(ws.title)


# ──────────────────────────── Парсинг блока ─────────────────────────────

@dataclass
class WellRecord:
    date: date
    ggu: str
    well: str
    choke_mm: Optional[float]
    p_wellhead: Optional[float]
    p_annular: Optional[float]
    annular_packer: bool
    p_flowline: Optional[float]
    q_gas_total: Optional[float]
    q_gas_working: Optional[float]
    shutdown_min: Optional[float]
    p_static: Optional[float]
    source_sheet: str


def _ggu_from_marker(value) -> str:
    """Нормализовать значение № ГСП ('2' → 'ГСП-2', 'БВН' → 'БВН')."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return f"ГСП-{int(value)}"
    s = str(value).strip()
    if not s:
        return ""
    if s.upper() == "БВН" or s.lower() == "бвн":
        return "БВН"
    # Если «2», «3» и пр.
    try:
        f = float(s.replace(",", "."))
        if f.is_integer():
            return f"ГСП-{int(f)}"
    except ValueError:
        pass
    return s


def parse_block(ws, header_row: int, sheet_date: date, sheet_name: str) -> list[WellRecord]:
    """Распарсить один блок (10 строк), начиная со строки header_row."""
    # Предварительно считаем ряды
    rows = {
        key: {c: ws.cell(row=header_row + off, column=c).value
              for c in range(MIN_COL, MAX_COL + 1)}
        for key, off in ROW_OFFSETS.items()
    }

    # Forward-fill значений ГСП по столбцам
    ggu_by_col: dict[int, str] = {}
    current = ""
    for c in range(MIN_COL, MAX_COL + 1):
        marker = _ggu_from_marker(rows["ggu"][c])
        if marker:
            current = marker
        ggu_by_col[c] = current

    records: list[WellRecord] = []
    for c in range(MIN_COL, MAX_COL + 1):
        well = _well_id(rows["well"][c])
        if not well:
            continue
        ggu = ggu_by_col.get(c, "")
        if not ggu:
            log.debug("%s: ячейка col=%d — № ГСП не определён, пропуск скв. %s",
                      sheet_name, c, well)
            continue

        raw_annular = rows["p_annular"][c]
        annular_packer = _is_packer(raw_annular)
        p_annular = None if annular_packer else _to_float(raw_annular)

        records.append(WellRecord(
            date=sheet_date,
            ggu=ggu,
            well=well,
            choke_mm=_to_float(rows["choke_mm"][c]),
            p_wellhead=_to_float(rows["p_wellhead"][c]),
            p_annular=p_annular,
            annular_packer=annular_packer,
            p_flowline=_to_float(rows["p_flowline"][c]),
            q_gas_total=_to_float(rows["q_gas_total"][c]),
            q_gas_working=_to_float(rows["q_gas_working"][c]),
            shutdown_min=_to_float(rows["shutdown_min"][c]),
            p_static=_to_float(rows["p_static"][c]),
            source_sheet=sheet_name,
        ))
    return records


# ──────────────────────────── Основной парсер ───────────────────────────

def parse_workbook(xlsx_path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    all_records: list[WellRecord] = []
    skipped: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_date = extract_date(ws)
        if sheet_date is None:
            log.warning("Пропускаю лист %r: не удалось определить дату", sheet_name)
            skipped.append(sheet_name)
            continue

        sheet_records: list[WellRecord] = []
        for hr in BLOCK_HEADER_ROWS:
            if hr > ws.max_row:
                continue
            sheet_records.extend(parse_block(ws, hr, sheet_date, sheet_name))

        log.info("Лист %r → %s: %d скважин", sheet_name, sheet_date, len(sheet_records))
        all_records.extend(sheet_records)

    if skipped:
        log.warning("Пропущено листов без даты: %d — %s", len(skipped), skipped)

    if not all_records:
        return pd.DataFrame()

    df = pd.DataFrame([r.__dict__ for r in all_records])

    # Обработка дублей (скважина × дата): если один и тот же ключ встречается
    # несколько раз (напр. дубликат листа '11,01,26' и '11,01,26 '),
    # оставляем последний по порядку и предупреждаем.
    dup_mask = df.duplicated(subset=["date", "well", "ggu"], keep="last")
    if dup_mask.any():
        dup_rows = df[df.duplicated(subset=["date", "well", "ggu"], keep=False)]
        log.warning(
            "Обнаружены дубликаты (скважина × дата), оставлена последняя запись:\n%s",
            dup_rows[["date", "ggu", "well", "source_sheet"]].to_string(index=False),
        )
        df = df.loc[~dup_mask].reset_index(drop=True)

    # Сортировка: дата → ГСП → скважина
    df = df.sort_values(["date", "ggu", "well"]).reset_index(drop=True)
    return df


# ───────────────────────────────── CLI ──────────────────────────────────

COLUMN_ORDER = [
    "date", "ggu", "well",
    "choke_mm",
    "p_wellhead", "p_annular", "annular_packer", "p_flowline",
    "q_gas_total", "q_gas_working",
    "shutdown_min", "p_static",
    "source_sheet",
]


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Парсер суточных сводок скважин → CSV")
    p.add_argument("input", type=Path, help="Путь к исходному .xlsx")
    p.add_argument("-o", "--output", type=Path, default=None,
                   help="Путь к результату (по умолчанию рядом с input, с расширением .csv)")
    p.add_argument("--sep", default=",", help="Разделитель CSV (по умолчанию ',')")
    p.add_argument("--encoding", default="utf-8-sig",
                   help="Кодировка файла (по умолчанию utf-8-sig для Excel)")
    p.add_argument("-v", "--verbose", action="store_true", help="Подробный лог")
    args = p.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    if not args.input.exists():
        log.error("Файл не найден: %s", args.input)
        return 2

    out = args.output or args.input.with_suffix(".csv")

    log.info("Читаю %s ...", args.input)
    df = parse_workbook(args.input)

    if df.empty:
        log.error("Не удалось извлечь ни одной записи.")
        return 1

    df = df[COLUMN_ORDER]
    df.to_csv(out, index=False, sep=args.sep, encoding=args.encoding)

    log.info(
        "Готово. Записано %d строк, %d уникальных скважин, %d дат → %s",
        len(df), df["well"].nunique(), df["date"].nunique(), out,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())